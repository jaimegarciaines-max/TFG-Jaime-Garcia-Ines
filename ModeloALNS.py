"""
Modelo híbrido exacto + ALNS con generación automática de instancias.

OBJETIVO DEL PROGRAMA
---------------------
El programa genera automáticamente instancias de un problema de asignación y
secuenciación de orders en líneas paralelas, las resuelve con un modelo exacto
de Gurobi y, si Gurobi encuentra una solución factible pero no demuestra
optimalidad, intenta mejorarla mediante una metaheurística ALNS.

El problema representa una planificación de producción:
- Hay varias líneas paralelas.
- Todas las orders deben asignarse obligatoriamente.
- Todas las orders pueden ir a cualquier línea.
- Cada línea tiene una capacidad máxima.
- La capacidad total se calcula mediante una holgura sobre las toneladas totales.
- Las orders asignadas a una misma línea deben secuenciarse.
- El coste de setup entre dos orders se toma directamente de una matriz de costes.
- El coste de inicio de cada línea también se toma directamente de una matriz.


ESTRUCTURA DEL CÓDIGO
---------------------
1. Configuración general del experimento.
2. Generación automática de instancias.
3. Guardado de inputs generados.
4. Funciones auxiliares de solución.
5. Operadores destroy del ALNS.
6. Operadores repair del ALNS.
7. ALNS.
8. Modelo exacto de Gurobi.
9. Output detallado por instancia.
10. Excel resumen.
11. Ejecución principal.
"""

import os
import random
import math
import time

import gurobipy as gp
from gurobipy import GRB

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# 1. CONFIGURACIÓN GENERAL DEL EXPERIMENTO
# ============================================================

CARPETA_RESULTADOS = "resultados_tuning_adaptativo"
CARPETA_INPUTS = os.path.join(CARPETA_RESULTADOS, "inputsAdaptar")
CARPETA_OUTPUTS = os.path.join(CARPETA_RESULTADOS, "outputsAdaptar")
EXCEL_RESUMEN = os.path.join(CARPETA_RESULTADOS, "resultados_tuning_adaptativo.xlsx")

USAR_SEMILLA =  True
SEED = 789
EMPAREJAR_INSTANCIAS_ENTRE_CONFIGURACIONES = True

GUROBI_OUTPUT_FLAG = 1
GUROBI_LOG_TO_CONSOLE = 1
PRINT_PROGRESS = True


CONFIGURACIONES = [
   {
        "nombre": "100_orders_adaptativo",
        "num_instancias": 5,
        "NUM_LINES": 4,
        "NUM_ORDERS": 100,
        "MEDIA_TONELADAS": 20,
        "DESV_TONELADAS": 4,
        "HOLGURA_CAPACIDAD": 0.10,
        "PROB_COSTE_CERO": 0.0005,
        "MIN_COSTE_SETUP": 1,
        "MAX_COSTE_SETUP": 100,
        "PROB_COSTE_INICIO_CERO": 0.0005,
        "MIN_COSTE_INICIO": 1,
        "MAX_COSTE_INICIO": 100,
        "MIN_TIME": 180,
        "MAX_TIME": 120,
        "TARGET_GAP": 0,
        "ALNS_MAX_ITER": 500000,
        "ALNS_MAX_TIME": 300,
        "ALNS_PORCENTAJE_REMOVAL": 0.10,
    },
    {
        "nombre": "120_orders_adaptativo",
        "num_instancias": 5,
        "NUM_LINES": 4,
        "NUM_ORDERS": 120,
        "MEDIA_TONELADAS": 20,
        "DESV_TONELADAS": 4,
        "HOLGURA_CAPACIDAD": 0.10,
        "PROB_COSTE_CERO": 0.0005,
        "MIN_COSTE_SETUP": 1,
        "MAX_COSTE_SETUP": 100,
        "PROB_COSTE_INICIO_CERO": 0.0005,
        "MIN_COSTE_INICIO": 1,
        "MAX_COSTE_INICIO": 100,
        "MIN_TIME": 180,
        "MAX_TIME": 120,
        "TARGET_GAP": 0,
        "ALNS_MAX_ITER": 500000,
        "ALNS_MAX_TIME": 300,
        "ALNS_PORCENTAJE_REMOVAL": 0.10,
    },
    {
        "nombre": "140_orders_adaptativo",
        "num_instancias": 5,
        "NUM_LINES": 4,
        "NUM_ORDERS": 140,
        "MEDIA_TONELADAS": 20,
        "DESV_TONELADAS": 4,
        "HOLGURA_CAPACIDAD": 0.10,
        "PROB_COSTE_CERO": 0.0005,
        "MIN_COSTE_SETUP": 1,
        "MAX_COSTE_SETUP": 100,
        "PROB_COSTE_INICIO_CERO": 0.0005,
        "MIN_COSTE_INICIO": 1,
        "MAX_COSTE_INICIO": 100,
        "MIN_TIME": 180,
        "MAX_TIME": 120,
        "TARGET_GAP": 0,
        "ALNS_MAX_ITER": 500000,
        "ALNS_MAX_TIME": 300,
        "ALNS_PORCENTAJE_REMOVAL": 0.10,
    },
]


def total_instancias_experimento():
    """Devuelve el número total de instancias que se resolverán."""
    return sum(config["num_instancias"] for config in CONFIGURACIONES)


# ============================================================
# 2. GENERACIÓN AUTOMÁTICA DE INSTANCIAS
# ============================================================

def preparar_carpetas():
    """Crea las carpetas donde se guardarán inputs, outputs y Excel."""
    os.makedirs(CARPETA_RESULTADOS, exist_ok=True)
    os.makedirs(CARPETA_INPUTS, exist_ok=True)
    os.makedirs(CARPETA_OUTPUTS, exist_ok=True)


def obtener_seed_instancia(indice_config, indice_local, instancia_id):
    """Calcula la semilla de una instancia según la configuración global."""
    if not USAR_SEMILLA:
        return None
    if EMPAREJAR_INSTANCIAS_ENTRE_CONFIGURACIONES:
        return SEED + indice_local
    return SEED + 100000 * indice_config + instancia_id


def generar_pesos(orders, config, rng):
    """Genera las toneladas de cada order con media y desviación configurables."""
    weight = {}
    media = config["MEDIA_TONELADAS"]
    desv = config["DESV_TONELADAS"]

    for i in orders:
        toneladas = round(abs(rng.normalvariate(media, desv)))
        if toneladas == 0:
            toneladas = 1
        weight[i] = toneladas

    return weight


def calcular_capacidades(weight, lines, config):
    """Calcula capacidad total y capacidad por línea usando holgura."""
    total_weight = sum(weight.values())
    total_capacity = total_weight * (1 + config["HOLGURA_CAPACIDAD"])
    capacity_line = total_capacity / len(lines)
    capacity = {l: capacity_line for l in lines}
    return total_weight, total_capacity, capacity


def generar_matriz_costes(orders, config, rng):
    """Genera una matriz asimétrica de costes entre orders."""
    coste = {}
    prob_cero = config["PROB_COSTE_CERO"]
    min_coste = config["MIN_COSTE_SETUP"]
    max_coste = config["MAX_COSTE_SETUP"]

    for i in orders:
        for j in orders:
            if i == j:
                coste[(i, j)] = 0

    for pos_i, i in enumerate(orders):
        for j in orders[pos_i + 1:]:
            if rng.random() <= prob_cero:
                coste[(i, j)] = 0
                coste[(j, i)] = 0
            else:
                coste[(i, j)] = rng.randint(min_coste, max_coste)
                coste[(j, i)] = rng.randint(min_coste, max_coste)

    return coste


def generar_costes_inicio(orders, lines, config, rng):
    """Genera el coste de iniciar cada línea con cada order."""
    coste_inicio = {}
    prob_cero = config["PROB_COSTE_INICIO_CERO"]
    min_coste = config["MIN_COSTE_INICIO"]
    max_coste = config["MAX_COSTE_INICIO"]

    for i in orders:
        for l in lines:
            if rng.random() <= prob_cero:
                coste_inicio[(i, l)] = 0
            else:
                coste_inicio[(i, l)] = rng.randint(min_coste, max_coste)

    return coste_inicio


def generar_instancia(instancia_id, indice_config, indice_local, config):
    """Genera todos los datos necesarios para una instancia."""
    seed_instancia = obtener_seed_instancia(indice_config, indice_local, instancia_id)
    rng = random.Random(seed_instancia) if seed_instancia is not None else random.Random()

    orders = list(range(config["NUM_ORDERS"]))
    lines = list(range(config["NUM_LINES"]))

    weight = generar_pesos(orders, config, rng)
    total_weight, total_capacity, capacity = calcular_capacidades(weight, lines, config)

    return {
        "instancia_id": instancia_id,
        "indice_config": indice_config,
        "indice_local": indice_local,
        "nombre_configuracion": config["nombre"],
        "num_orders": config["NUM_ORDERS"],
        "num_lines": config["NUM_LINES"],
        "orders": orders,
        "lines": lines,
        "weight": weight,
        "total_weight": total_weight,
        "total_capacity": total_capacity,
        "capacity": capacity,
        "coste": generar_matriz_costes(orders, config, rng),
        "coste_inicio": generar_costes_inicio(orders, lines, config, rng),
        "media_toneladas": config["MEDIA_TONELADAS"],
        "desv_toneladas": config["DESV_TONELADAS"],
        "holgura_capacidad": config["HOLGURA_CAPACIDAD"],
        "prob_coste_cero": config["PROB_COSTE_CERO"],
        "min_coste_setup": config["MIN_COSTE_SETUP"],
        "max_coste_setup": config["MAX_COSTE_SETUP"],
        "prob_coste_inicio_cero": config["PROB_COSTE_INICIO_CERO"],
        "min_coste_inicio": config["MIN_COSTE_INICIO"],
        "max_coste_inicio": config["MAX_COSTE_INICIO"],
        "min_time": config["MIN_TIME"],
        "max_time": config["MAX_TIME"],
        "target_gap": config["TARGET_GAP"],
        "alns_max_iter": config["ALNS_MAX_ITER"],
        "alns_max_time": config["ALNS_MAX_TIME"],
        "alns_porcentaje_removal": config["ALNS_PORCENTAJE_REMOVAL"],
        "use_seed": USAR_SEMILLA,
        "seed_base": SEED if USAR_SEMILLA else "NO_SEED",
        "seed_instancia": seed_instancia if seed_instancia is not None else "NO_SEED",
        "match_instances_between_configs": EMPAREJAR_INSTANCIAS_ENTRE_CONFIGURACIONES,
    }


# ============================================================
# 3. GUARDADO DE INPUTS GENERADOS
# ============================================================

def ruta_input(instancia_id):
    return os.path.join(CARPETA_INPUTS, f"input_{instancia_id:03d}.txt")


def ruta_output(instancia_id):
    return os.path.join(CARPETA_OUTPUTS, f"output_{instancia_id:03d}.txt")


def guardar_input_generado(instancia):
    """Guarda en texto la instancia generada para poder auditarla."""
    orders = instancia["orders"]
    lines = instancia["lines"]
    weight = instancia["weight"]
    capacity = instancia["capacity"]
    coste = instancia["coste"]
    coste_inicio = instancia["coste_inicio"]

    with open(ruta_input(instancia["instancia_id"]), "w", encoding="utf-8") as f:
        f.write(f"# INPUT GENERADO AUTOMÁTICAMENTE - INSTANCIA {instancia['instancia_id']}\n")
        f.write(f"# CONFIGURACIÓN {instancia['nombre_configuracion']}\n")
        f.write(f"# ÍNDICE_LOCAL {instancia['indice_local']}\n")
        f.write(f"# USAR_SEMILLA {instancia['use_seed']}\n")
        f.write(f"# SEMILLA_BASE {instancia['seed_base']}\n")
        f.write(f"# SEMILLA_INSTANCIA {instancia['seed_instancia']}\n")
        f.write(f"# HOLGURA_CAPACIDAD {instancia['holgura_capacidad']}\n")
        f.write(f"# TOTAL_WEIGHT {instancia['total_weight']}\n")
        f.write(f"# TOTAL_CAPACITY {instancia['total_capacity']}\n\n")

        f.write(f"NUM_ORDERS {instancia['num_orders']}\n")
        f.write(f"NUM_LINES {instancia['num_lines']}\n\n")
        f.write(f"TOTAL_WEIGHT {instancia['total_weight']}\n")
        f.write(f"TOTAL_CAPACITY {instancia['total_capacity']}\n\n")

        f.write("WEIGHT\n")
        for i in orders:
            f.write(f"{i} {weight[i]}\n")
        f.write("\n")

        f.write("CAPACITY\n")
        for l in lines:
            f.write(f"{l} {capacity[l]}\n")
        f.write("\n")

        f.write("COST_MATRIX\n")
        f.write("order " + " ".join(str(j) for j in orders) + "\n")
        for i in orders:
            fila = [str(coste[(i, j)]) for j in orders]
            f.write(f"{i} " + " ".join(fila) + "\n")
        f.write("\n")

        f.write("START_COST\n")
        f.write("order " + " ".join(f"line_{l}" for l in lines) + "\n")
        for i in orders:
            fila = [str(coste_inicio[(i, l)]) for l in lines]
            f.write(f"{i} " + " ".join(fila) + "\n")


# ============================================================
# 4. FUNCIONES AUXILIARES DE SOLUCIÓN
# ============================================================

I = "inicio"
F = "final"


def extraer_solucion_desde_gurobi(x, orders, lines):
    """Reconstruye las secuencias de cada línea a partir de los arcos activos x."""
    solucion = {}

    for l in lines:
        inicio = None
        for j in orders:
            if x[I, j, l].x > 0.5:
                inicio = j
                break

        if inicio is None:
            solucion[l] = []
            continue

        secuencia = [inicio]
        actual = inicio

        while True:
            if x[actual, F, l].x > 0.5:
                break

            siguiente = None
            for j in orders:
                if j != actual and x[actual, j, l].x > 0.5:
                    siguiente = j
                    break

            if siguiente is None:
                break

            secuencia.append(siguiente)
            actual = siguiente

        solucion[l] = secuencia

    return solucion


def pedidos_asignados(solucion):
    asignados = []
    for secuencia in solucion.values():
        asignados.extend(secuencia)
    return asignados


def carga_solucion(solucion, weight):
    return {l: sum(weight[i] for i in secuencia) for l, secuencia in solucion.items()}


def coste_solucion(solucion, coste, coste_inicio):
    total = 0
    for l, secuencia in solucion.items():
        if len(secuencia) == 0:
            continue
        total += coste_inicio[(secuencia[0], l)]
        for pos in range(len(secuencia) - 1):
            i = secuencia[pos]
            j = secuencia[pos + 1]
            total += coste[(i, j)]
    return total


def es_factible(solucion, orders, weight, capacity):
    asignados = pedidos_asignados(solucion)
    if sorted(asignados) != sorted(orders):
        return False
    if len(asignados) != len(set(asignados)):
        return False
    for l, carga_l in carga_solucion(solucion, weight).items():
        if carga_l > capacity[l] + 1e-9:
            return False
    return True


def elegir_por_pesos(pesos):
    total = sum(pesos.values())
    if total <= 0:
        return random.choice(list(pesos.keys()))
    r = random.uniform(0, total)
    acumulado = 0
    for nombre, peso in pesos.items():
        acumulado += peso
        if r <= acumulado:
            return nombre
    return list(pesos.keys())[-1]


# ============================================================
# 5. ALNS: OPERADORES DESTROY
# ============================================================

def worst_removal(solucion, q, coste, coste_inicio):
    coste_base = coste_solucion(solucion, coste, coste_inicio)
    contribuciones = []
    for l, secuencia in solucion.items():
        for pedido in secuencia:
            candidata = {ll: ss.copy() for ll, ss in solucion.items()}
            candidata[l].remove(pedido)
            nuevo_coste = coste_solucion(candidata, coste, coste_inicio)
            ahorro = coste_base - nuevo_coste
            contribuciones.append((ahorro, l, pedido))
    contribuciones.sort(reverse=True)
    nueva_solucion = {l: sec.copy() for l, sec in solucion.items()}
    eliminados = []
    for _, l, pedido in contribuciones[:q]:
        if pedido in nueva_solucion[l]:
            nueva_solucion[l].remove(pedido)
            eliminados.append(pedido)
    return nueva_solucion, eliminados


def random_removal(solucion, q, coste, coste_inicio):
    todos = []
    for l, secuencia in solucion.items():
        for pedido in secuencia:
            todos.append((l, pedido))
    q_real = min(q, len(todos))
    if q_real == 0:
        return {l: sec.copy() for l, sec in solucion.items()}, []
    seleccionados = random.sample(todos, q_real)
    nueva_solucion = {l: sec.copy() for l, sec in solucion.items()}
    eliminados = []
    for l, pedido in seleccionados:
        if pedido in nueva_solucion[l]:
            nueva_solucion[l].remove(pedido)
            eliminados.append(pedido)
    return nueva_solucion, eliminados


def related_removal(solucion, q, coste, coste_inicio):
    todos = []
    for secuencia in solucion.values():
        todos.extend(secuencia)
    if len(todos) == 0:
        return {l: sec.copy() for l, sec in solucion.items()}, []
    q_real = min(q, len(todos))
    semilla = random.choice(todos)
    eliminados_set = {semilla}
    while len(eliminados_set) < q_real:
        referencia = random.choice(list(eliminados_set))
        candidatos = [p for p in todos if p not in eliminados_set]
        if not candidatos:
            break
        candidatos.sort(key=lambda p: min(coste[(referencia, p)], coste[(p, referencia)]))
        limite = max(1, min(5, len(candidatos)))
        elegido = random.choice(candidatos[:limite])
        eliminados_set.add(elegido)
    nueva_solucion = {l: sec.copy() for l, sec in solucion.items()}
    eliminados = []
    for pedido in eliminados_set:
        for l in nueva_solucion:
            if pedido in nueva_solucion[l]:
                nueva_solucion[l].remove(pedido)
                eliminados.append(pedido)
                break
    return nueva_solucion, eliminados


# ============================================================
# 6. ALNS: OPERADORES REPAIR
# ============================================================

def mejor_insercion_pedido(solucion, pedido, lines, coste, coste_inicio, weight, capacity):
    opciones = []
    for l in lines:
        carga_actual = sum(weight[i] for i in solucion[l])
        if carga_actual + weight[pedido] > capacity[l] + 1e-9:
            continue
        for pos in range(len(solucion[l]) + 1):
            candidata = {ll: ss.copy() for ll, ss in solucion.items()}
            candidata[l].insert(pos, pedido)
            c = coste_solucion(candidata, coste, coste_inicio)
            opciones.append((c, l, pos))
    opciones.sort(key=lambda x: x[0])
    return opciones


def cheapest_insertion(solucion_parcial, eliminados, lines, coste, coste_inicio, weight, capacity):
    solucion = {l: sec.copy() for l, sec in solucion_parcial.items()}
    for pedido in eliminados:
        opciones = mejor_insercion_pedido(solucion, pedido, lines, coste, coste_inicio, weight, capacity)
        if len(opciones) == 0:
            raise ValueError(f"No se puede reinsertar el pedido {pedido}")
        _, mejor_linea, mejor_pos = opciones[0]
        solucion[mejor_linea].insert(mejor_pos, pedido)
    return solucion


def regret_2_insertion(solucion_parcial, eliminados, lines, coste, coste_inicio, weight, capacity):
    solucion = {l: sec.copy() for l, sec in solucion_parcial.items()}
    pendientes = eliminados.copy()
    while pendientes:
        mejor_pedido = None
        mejor_linea = None
        mejor_pos = None
        mejor_regret = -float("inf")
        mejor_coste = float("inf")
        for pedido in pendientes:
            opciones = mejor_insercion_pedido(solucion, pedido, lines, coste, coste_inicio, weight, capacity)
            if len(opciones) == 0:
                continue
            coste_1, linea_1, pos_1 = opciones[0]
            regret = opciones[1][0] - coste_1 if len(opciones) >= 2 else float("inf")
            if regret > mejor_regret or (regret == mejor_regret and coste_1 < mejor_coste):
                mejor_regret = regret
                mejor_coste = coste_1
                mejor_pedido = pedido
                mejor_linea = linea_1
                mejor_pos = pos_1
        if mejor_pedido is None:
            raise ValueError("No se ha podido reinsertar alguno de los pedidos eliminados")
        solucion[mejor_linea].insert(mejor_pos, mejor_pedido)
        pendientes.remove(mejor_pedido)
    return solucion


# ============================================================
# 7. ALNS
# ============================================================

def alns(solucion_inicial, orders, lines, coste, coste_inicio, weight, capacity,
         max_iter=500000, max_time=300, porcentaje_removal=0.20):
    tiempo_inicio = time.time()
    actual = {l: sec.copy() for l, sec in solucion_inicial.items()}
    mejor = {l: sec.copy() for l, sec in solucion_inicial.items()}
    coste_actual = coste_solucion(actual, coste, coste_inicio)
    mejor_coste = coste_actual
    q = max(1, int(len(orders) * porcentaje_removal))
    temperatura = max(1, 0.05 * coste_actual)
    enfriamiento = 0.995

    removal_operators = {
        "worst_removal": worst_removal,
        "random_removal": random_removal,
        "related_removal": related_removal,
    }
    repair_operators = {
        "cheapest_insertion": cheapest_insertion,
        "regret_2_insertion": regret_2_insertion,
    }
    pesos_removal = {nombre: 1.0 for nombre in removal_operators}
    pesos_repair = {nombre: 1.0 for nombre in repair_operators}
    puntuacion_removal = {nombre: 0.0 for nombre in removal_operators}
    puntuacion_repair = {nombre: 0.0 for nombre in repair_operators}
    usos_removal = {nombre: 0 for nombre in removal_operators}
    usos_repair = {nombre: 0 for nombre in repair_operators}

    segmento = 10
    factor_reaccion = 0.20
    premio_mejor_global = 10
    premio_mejora_actual = 5
    premio_aceptada_peor = 1
    iteraciones = 0

    for _ in range(max_iter):
        if time.time() - tiempo_inicio >= max_time:
            break
        iteraciones += 1
        nombre_removal = elegir_por_pesos(pesos_removal)
        nombre_repair = elegir_por_pesos(pesos_repair)
        usos_removal[nombre_removal] += 1
        usos_repair[nombre_repair] += 1
        parcial, eliminados = removal_operators[nombre_removal](actual, q, coste, coste_inicio)
        try:
            candidata = repair_operators[nombre_repair](parcial, eliminados, lines, coste, coste_inicio, weight, capacity)
        except ValueError:
            continue
        if not es_factible(candidata, orders, weight, capacity):
            continue
        coste_candidata = coste_solucion(candidata, coste, coste_inicio)
        diferencia = coste_candidata - coste_actual
        aceptar = False
        premio = 0
        if coste_candidata < mejor_coste:
            aceptar = True
            premio = premio_mejor_global
        elif diferencia < 0:
            aceptar = True
            premio = premio_mejora_actual
        else:
            probabilidad = math.exp(-diferencia / temperatura) if temperatura > 1e-9 else 0
            aceptar = random.random() < probabilidad
            if aceptar:
                premio = premio_aceptada_peor
        if premio > 0:
            puntuacion_removal[nombre_removal] += premio
            puntuacion_repair[nombre_repair] += premio
        if aceptar:
            actual = candidata
            coste_actual = coste_candidata
        if coste_candidata < mejor_coste:
            mejor = candidata
            mejor_coste = coste_candidata
        temperatura *= enfriamiento
        if iteraciones % segmento == 0:
            for nombre in pesos_removal:
                if usos_removal[nombre] > 0:
                    rendimiento = puntuacion_removal[nombre] / usos_removal[nombre]
                    pesos_removal[nombre] = (1 - factor_reaccion) * pesos_removal[nombre] + factor_reaccion * rendimiento
                puntuacion_removal[nombre] = 0.0
                usos_removal[nombre] = 0
                pesos_removal[nombre] = max(0.05, pesos_removal[nombre])
            for nombre in pesos_repair:
                if usos_repair[nombre] > 0:
                    rendimiento = puntuacion_repair[nombre] / usos_repair[nombre]
                    pesos_repair[nombre] = (1 - factor_reaccion) * pesos_repair[nombre] + factor_reaccion * rendimiento
                puntuacion_repair[nombre] = 0.0
                usos_repair[nombre] = 0
                pesos_repair[nombre] = max(0.05, pesos_repair[nombre])

    estadisticas = {
        "iteraciones": iteraciones,
        "tiempo": time.time() - tiempo_inicio,
        "coste_inicial": coste_solucion(solucion_inicial, coste, coste_inicio),
        "mejor_coste": mejor_coste,
        "pesos_removal": pesos_removal,
        "pesos_repair": pesos_repair,
    }
    return mejor, mejor_coste, estadisticas


# ============================================================
# 8. MODELO EXACTO DE GUROBI
# ============================================================

def resolver_instancia(instancia, total_instancias):
    orders = instancia["orders"]
    lines = instancia["lines"]
    weight = instancia["weight"]
    capacity = instancia["capacity"]
    coste = instancia["coste"]
    coste_inicio = instancia["coste_inicio"]
    instancia_id = instancia["instancia_id"]

    m = gp.Model(f"modelo_matriz_costes_instancia_{instancia_id:03d}")
    m.setParam("OutputFlag", GUROBI_OUTPUT_FLAG)
    m.setParam("LogToConsole", GUROBI_LOG_TO_CONSOLE)
    m._parada_por_gap_y_tiempo = False
    m._parada_por_tiempo_maximo = False

    def parada_personalizada(model, where):
        if where == GRB.Callback.MIP:
            runtime = model.cbGet(GRB.Callback.RUNTIME)
            if runtime >= instancia["max_time"]:
                model._parada_por_tiempo_maximo = True
                model.terminate()
                return
            obj_best = model.cbGet(GRB.Callback.MIP_OBJBST)
            obj_bound = model.cbGet(GRB.Callback.MIP_OBJBND)
            if obj_best < GRB.INFINITY and abs(obj_best) > 1e-9:
                gap = abs(obj_best - obj_bound) / abs(obj_best)
                if runtime >= instancia["min_time"] and gap <= instancia["target_gap"]:
                    model._parada_por_gap_y_tiempo = True
                    model.terminate()

    arcos_orders = [(i, j, l) for i in orders for j in orders if i != j for l in lines]
    arcos_inicio = [(I, j, l) for j in orders for l in lines]
    arcos_final = [(i, F, l) for i in orders for l in lines]
    arcos = arcos_orders + arcos_inicio + arcos_final

    x = m.addVars(arcos, vtype=GRB.BINARY, name="x")
    u = m.addVars(orders, lines, vtype=GRB.CONTINUOUS, lb=0, ub=len(orders), name="u")
    y = m.addVars(orders, lines, vtype=GRB.BINARY, name="assign")
    carga = m.addVars(lines, vtype=GRB.CONTINUOUS, lb=0, name="load")

    m.setObjective(
        gp.quicksum(coste[(i, j)] * x[i, j, l] for i in orders for j in orders if i != j for l in lines)
        + gp.quicksum(coste_inicio[(j, l)] * x[I, j, l] for j in orders for l in lines),
        GRB.MINIMIZE,
    )

    for i in orders:
        m.addConstr(
            gp.quicksum(x[i, j, l] for j in orders if i != j for l in lines)
            + gp.quicksum(x[i, F, l] for l in lines)
            == 1,
            name=f"salida_unica_{i}",
        )

    for j in orders:
        m.addConstr(
            gp.quicksum(x[i, j, l] for i in orders if i != j for l in lines)
            + gp.quicksum(x[I, j, l] for l in lines)
            == 1,
            name=f"entrada_unica_{j}",
        )

    for l in lines:
        for k in orders:
            m.addConstr(
                gp.quicksum(x[i, k, l] for i in orders if i != k) + x[I, k, l]
                == gp.quicksum(x[k, j, l] for j in orders if j != k) + x[k, F, l],
                name=f"flujo_linea_{l}_order_{k}",
            )

    for l in lines:
        m.addConstr(gp.quicksum(x[I, j, l] for j in orders) <= 1, name=f"un_inicio_linea_{l}")
        m.addConstr(gp.quicksum(x[i, F, l] for i in orders) <= 1, name=f"un_final_linea_{l}")

    for l in lines:
        for i in orders:
            m.addConstr(
                y[i, l] == gp.quicksum(x[h, i, l] for h in orders if h != i) + x[I, i, l],
                name=f"def_assign_{i}_{l}",
            )

    for l in lines:
        m.addConstr(carga[l] == gp.quicksum(weight[i] * y[i, l] for i in orders), name=f"def_load_{l}")
        m.addConstr(carga[l] <= capacity[l], name=f"capacidad_linea_{l}")

    n = len(orders)
    for i in orders:
        for j in orders:
            if i != j:
                for l in lines:
                    m.addConstr(u[i, l] - u[j, l] + n * x[i, j, l] <= n - 1, name=f"mtz_{i}_{j}_{l}")

    if PRINT_PROGRESS:
        print("\n==================================================")
        print(f"RESOLVIENDO INSTANCIA {instancia_id}/{total_instancias}")
        print(f"Configuración: {instancia['nombre_configuracion']} | local {instancia['indice_local']}")
        print("==================================================\n")

    m.optimize(parada_personalizada)

    resultado = {
        "instancia_id": instancia_id,
        "status_gurobi": m.status,
        "solucion_gurobi": None,
        "solucion_alns": None,
        "alns_ejecutado": False,
        "stats_alns": None,
        "mip_gap": None,
        "runtime_gurobi": getattr(m, "Runtime", None),
        "obj_gurobi": None,
        "obj_alns": None,
        "parada_por_gap_y_tiempo": m._parada_por_gap_y_tiempo,
        "parada_por_tiempo_maximo": m._parada_por_tiempo_maximo,
        "sol_count": m.SolCount,
    }

    if m.SolCount > 0:
        solucion_gurobi = extraer_solucion_desde_gurobi(x, orders, lines)
        resultado["solucion_gurobi"] = solucion_gurobi
        resultado["obj_gurobi"] = m.objVal
        try:
            resultado["mip_gap"] = m.MIPGap
        except Exception:
            resultado["mip_gap"] = None

        if m.status != GRB.OPTIMAL:
            if USAR_SEMILLA and instancia["seed_instancia"] != "NO_SEED":
                random.seed(instancia["seed_instancia"] + 999999)
            solucion_alns, coste_alns, stats_alns = alns(
                solucion_gurobi,
                orders,
                lines,
                coste,
                coste_inicio,
                weight,
                capacity,
                max_iter=instancia["alns_max_iter"],
                max_time=instancia["alns_max_time"],
                porcentaje_removal=instancia["alns_porcentaje_removal"],
            )
            resultado["solucion_alns"] = solucion_alns
            resultado["obj_alns"] = coste_alns
            resultado["alns_ejecutado"] = True
            resultado["stats_alns"] = stats_alns

    return resultado


# ============================================================
# 9. OUTPUT DETALLADO DE CADA INSTANCIA
# ============================================================

def estado_gurobi_excel(resultado):
    if resultado["sol_count"] == 0:
        return "Sin solución factible"
    if resultado["status_gurobi"] == GRB.OPTIMAL:
        return "Solución óptima"
    return "Solución factible"


def motivo_parada_output(resultado):
    if resultado["sol_count"] == 0:
        return "Sin solución factible"
    if resultado["status_gurobi"] == GRB.OPTIMAL:
        return "Parada por solución óptima"
    if resultado["parada_por_tiempo_maximo"]:
        return "Parada por tiempo máximo alcanzado"
    if resultado["parada_por_gap_y_tiempo"]:
        return "Parada por gap objetivo y tiempo mínimo alcanzados"
    return "Parada sin prueba de optimalidad"


def resultado_final_excel(resultado):
    if resultado["sol_count"] == 0:
        return "Sin solución factible"
    if resultado["status_gurobi"] == GRB.OPTIMAL:
        return "Solución óptima"
    if resultado["alns_ejecutado"] and resultado["stats_alns"] is not None:
        mejora = resultado["stats_alns"]["coste_inicial"] - resultado["stats_alns"]["mejor_coste"]
        if mejora > 1e-9:
            return "Solución factible con mejora de ALNS"
        return "Solución factible sin mejora de ALNS"
    return "Solución factible"


def escribir(linea, f):
    print(linea)
    f.write(linea + "\n")


def escribir_resumen_solucion(nombre, solucion, instancia, f):
    orders = instancia["orders"]
    lines = instancia["lines"]
    weight = instancia["weight"]
    capacity = instancia["capacity"]
    coste = instancia["coste"]
    coste_inicio = instancia["coste_inicio"]
    cargas = carga_solucion(solucion, weight)
    coste_setup = coste_solucion(solucion, coste, coste_inicio)
    factible = es_factible(solucion, orders, weight, capacity)

    escribir(f"Resumen {nombre}:", f)
    escribir(f"Factible: {'Sí' if factible else 'No'}", f)
    escribir(f"Toneladas totales: {instancia['total_weight']}", f)
    escribir(f"Capacidad total: {instancia['total_capacity']}", f)
    escribir(f"Holgura capacidad: {instancia['holgura_capacidad']}", f)
    escribir(f"Número de orders asignadas: {len(pedidos_asignados(solucion))}", f)
    escribir(f"Coste setup: {coste_setup}", f)
    escribir("", f)

    escribir(f"Carga por línea según {nombre}:", f)
    for l in lines:
        escribir(f"Línea {l}: {cargas[l]} toneladas / capacidad {capacity[l]}", f)

    escribir("", f)
    escribir(f"Secuencias según {nombre}:", f)
    for l in lines:
        escribir(f"\nLínea {l}:", f)
        secuencia = solucion[l]
        if len(secuencia) == 0:
            escribir("  Línea vacía", f)
        else:
            escribir("  start -> " + " -> ".join(map(str, secuencia)) + " -> end", f)
    escribir("", f)


def guardar_output_instancia(instancia, resultado):
    instancia_id = instancia["instancia_id"]
    with open(ruta_output(instancia_id), "w", encoding="utf-8") as f:
        escribir("==================================================", f)
        escribir("MODELO HÍBRIDO CON MATRIZ DE COSTES", f)
        escribir("==================================================", f)
        escribir(f"Instancia: {instancia_id}", f)
        escribir(f"Configuración: {instancia['nombre_configuracion']}", f)
        escribir(f"Índice local: {instancia['indice_local']}", f)
        escribir(f"Archivo input generado: {ruta_input(instancia_id)}", f)
        escribir(f"Archivo output generado: {ruta_output(instancia_id)}", f)
        escribir(f"Excel resumen: {EXCEL_RESUMEN}", f)
        escribir("", f)

        escribir("Parámetros de la instancia:", f)
        escribir(f"Número de líneas: {instancia['num_lines']}", f)
        escribir(f"Número de orders: {instancia['num_orders']}", f)
        escribir(f"Toneladas totales: {instancia['total_weight']}", f)
        escribir(f"Capacidad total: {instancia['total_capacity']}", f)
        escribir(f"Holgura capacidad: {instancia['holgura_capacidad']}", f)
        escribir(f"Probabilidad coste cero entre orders: {instancia['prob_coste_cero']}", f)
        escribir(f"Probabilidad coste inicio cero: {instancia['prob_coste_inicio_cero']}", f)
        escribir(f"USAR_SEMILLA: {instancia['use_seed']}", f)
        escribir(f"SEMILLA_BASE: {instancia['seed_base']}", f)
        escribir(f"SEMILLA_INSTANCIA: {instancia['seed_instancia']}", f)
        escribir(f"EMPAREJAR_INSTANCIAS_ENTRE_CONFIGURACIONES: {instancia['match_instances_between_configs']}", f)
        escribir("", f)

        escribir("Resultado Gurobi:", f)
        escribir(f"Estado simplificado: {estado_gurobi_excel(resultado)}", f)
        escribir(f"Motivo de parada: {motivo_parada_output(resultado)}", f)
        escribir(f"Tiempo Gurobi: {resultado['runtime_gurobi']} segundos", f)
        escribir(f"Valor objetivo Gurobi: {resultado['obj_gurobi']}", f)
        escribir(f"Gap final: {resultado['mip_gap'] if resultado['mip_gap'] is not None else 'No disponible'}", f)
        escribir(f"Resultado final: {resultado_final_excel(resultado)}", f)
        escribir("", f)

        if resultado["solucion_gurobi"] is not None:
            escribir_resumen_solucion("Gurobi", resultado["solucion_gurobi"], instancia, f)
            if resultado["alns_ejecutado"]:
                stats = resultado["stats_alns"]
                escribir("==================================================", f)
                escribir("MEJORA MEDIANTE ALNS", f)
                escribir("==================================================", f)
                escribir(f"Iteraciones ALNS realizadas: {stats['iteraciones']}", f)
                escribir(f"Tiempo ALNS: {stats['tiempo']} segundos", f)
                escribir(f"Coste inicial tomado de Gurobi: {stats['coste_inicial']}", f)
                escribir(f"Mejor coste encontrado por ALNS: {stats['mejor_coste']}", f)
                mejora = stats["coste_inicial"] - stats["mejor_coste"]
                porcentaje_mejora = 100 * mejora / stats["coste_inicial"] if abs(stats["coste_inicial"]) > 1e-9 else 0
                escribir(f"Mejora absoluta ALNS: {mejora}", f)
                escribir(f"Mejora porcentual ALNS: {porcentaje_mejora} %", f)
                escribir("", f)
                escribir_resumen_solucion("ALNS", resultado["solucion_alns"], instancia, f)
            else:
                escribir("==================================================", f)
                escribir("ALNS NO EJECUTADO", f)
                escribir("==================================================", f)
                if resultado["status_gurobi"] == GRB.OPTIMAL:
                    escribir("Gurobi ha demostrado optimalidad, por lo que no se aplica la metaheurística.", f)
                else:
                    escribir("No se ha ejecutado ALNS porque no existe solución factible inicial.", f)
        else:
            escribir("No hay solución factible que resumir.", f)


# ============================================================
# 10. EXCEL RESUMEN
# ============================================================

COLUMNAS_EXCEL = [
    "instancia_id", "indice_local", "num_orders", "num_lines",
    "toneladas_totales", "holgura_capacidad", "capacidad_total", "capacidad_por_linea",
    "media_toneladas", "desv_toneladas", "prob_coste_cero", "prob_coste_inicio_cero",
    "use_seed", "seed_base", "seed_instancia", "match_instances_between_configs",
    "estado_gurobi", "obj_gurobi", "gap_gurobi", "tiempo_gurobi",
    "alns_ejecutado", "obj_alns", "tiempo_alns", "iteraciones_alns",
    "coste_inicial_alns", "mejor_coste_alns", "mejora_absoluta_alns", "mejora_porcentual_alns",
    "resultado_final", "input_file", "output_file"
]


CABECERAS_EXCEL = [
    "id_instancia", "indice_local", "numero_orders", "numero_lineas",
    "toneladas_totales", "holgura_capacidad", "capacidad_total", "capacidad_por_linea",
    "media_toneladas", "desviacion_toneladas", "probabilidad_coste_cero", "probabilidad_coste_inicio_cero",
    "usar_semilla", "semilla_base", "semilla_instancia", "emparejar_instancias_entre_configuraciones",
    "estado_gurobi", "objetivo_gurobi", "gap_gurobi", "tiempo_gurobi",
    "alns_ejecutado", "objetivo_alns", "tiempo_alns", "iteraciones_alns",
    "coste_inicial_alns", "mejor_coste_alns", "mejora_absoluta_alns", "mejora_porcentual_alns",
    "resultado_final", "archivo_input", "archivo_output"
]

COLORES_RESULTADO_FINAL = {
    "Solución óptima": "D9EAD3",
    "Solución factible con mejora de ALNS": "D9EAF7",
    "Solución factible sin mejora de ALNS": "FFF2CC",
    "Sin solución factible": "F4CCCC",
    "Solución factible": "EADCF8",
}


def inicializar_excel():
    if os.path.exists(EXCEL_RESUMEN):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append(CABECERAS_EXCEL)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(EXCEL_RESUMEN)


def construir_fila_excel(instancia, resultado):
    fila = {col: None for col in COLUMNAS_EXCEL}
    fila["instancia_id"] = instancia["instancia_id"]
    fila["indice_local"] = instancia["indice_local"]
    fila["num_orders"] = instancia["num_orders"]
    fila["num_lines"] = instancia["num_lines"]
    fila["toneladas_totales"] = instancia["total_weight"]
    fila["holgura_capacidad"] = instancia["holgura_capacidad"]
    fila["capacidad_total"] = instancia["total_capacity"]
    fila["capacidad_por_linea"] = instancia["total_capacity"] / instancia["num_lines"]
    fila["media_toneladas"] = instancia["media_toneladas"]
    fila["desv_toneladas"] = instancia["desv_toneladas"]
    fila["prob_coste_cero"] = instancia["prob_coste_cero"]
    fila["prob_coste_inicio_cero"] = instancia["prob_coste_inicio_cero"]
    fila["use_seed"] = instancia["use_seed"]
    fila["seed_base"] = instancia["seed_base"]
    fila["seed_instancia"] = instancia["seed_instancia"]
    fila["match_instances_between_configs"] = instancia["match_instances_between_configs"]
    fila["estado_gurobi"] = estado_gurobi_excel(resultado)
    fila["obj_gurobi"] = resultado["obj_gurobi"]
    fila["gap_gurobi"] = resultado["mip_gap"]
    fila["tiempo_gurobi"] = resultado["runtime_gurobi"]
    fila["alns_ejecutado"] = resultado["alns_ejecutado"]
    fila["obj_alns"] = resultado["obj_alns"]
    fila["resultado_final"] = resultado_final_excel(resultado)
    fila["input_file"] = ruta_input(instancia["instancia_id"])
    fila["output_file"] = ruta_output(instancia["instancia_id"])
    if resultado["stats_alns"] is not None:
        stats = resultado["stats_alns"]
        mejora = stats["coste_inicial"] - stats["mejor_coste"]
        fila["tiempo_alns"] = stats["tiempo"]
        fila["iteraciones_alns"] = stats["iteraciones"]
        fila["coste_inicial_alns"] = stats["coste_inicial"]
        fila["mejor_coste_alns"] = stats["mejor_coste"]
        fila["mejora_absoluta_alns"] = mejora
        fila["mejora_porcentual_alns"] = 100 * mejora / stats["coste_inicial"] if abs(stats["coste_inicial"]) > 1e-9 else 0
    return [fila[col] for col in COLUMNAS_EXCEL]


def colorear_resultado_final(ws, fila_excel):
    col_idx = COLUMNAS_EXCEL.index("resultado_final") + 1
    celda = ws.cell(row=fila_excel, column=col_idx)
    color = COLORES_RESULTADO_FINAL.get(celda.value)
    if color is not None:
        celda.fill = PatternFill("solid", fgColor=color)
    celda.font = Font(bold=True)
    celda.alignment = Alignment(horizontal="center")


def actualizar_excel(instancia, resultado):
    inicializar_excel()
    wb = load_workbook(EXCEL_RESUMEN)
    ws = wb["Resultados"]
    ws.append(construir_fila_excel(instancia, resultado))
    fila_excel = ws.max_row
    colorear_resultado_final(ws, fila_excel)
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)
    wb.save(EXCEL_RESUMEN)


# ============================================================
# 11. EJECUCIÓN PRINCIPAL
# ============================================================

def main():
    preparar_carpetas()
    inicializar_excel()
    total_instancias = total_instancias_experimento()
    instancia_id = 0

    for indice_config, config in enumerate(CONFIGURACIONES, start=1):
        for indice_local in range(1, config["num_instancias"] + 1):
            instancia_id += 1
            instancia = generar_instancia(instancia_id, indice_config, indice_local, config)
            guardar_input_generado(instancia)
            resultado = resolver_instancia(instancia, total_instancias)
            guardar_output_instancia(instancia, resultado)
            actualizar_excel(instancia, resultado)
            if PRINT_PROGRESS:
                print("\n--------------------------------------------------")
                print(f"Instancia {instancia_id}/{total_instancias} completada")
                print(f"Configuración: {instancia['nombre_configuracion']} | local {indice_local}/{config['num_instancias']}")
                print(f"Input: {ruta_input(instancia_id)}")
                print(f"Output: {ruta_output(instancia_id)}")
                print(f"Excel: {EXCEL_RESUMEN}")
                print("--------------------------------------------------\n")


if __name__ == "__main__":
    main()
