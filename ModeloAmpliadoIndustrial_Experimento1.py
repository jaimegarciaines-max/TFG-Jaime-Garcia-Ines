"""
Modelo ampliado industrial para secuenciar orders en líneas paralelas.

OBJETIVO DEL PROGRAMA
---------------------
El programa genera automáticamente instancias del problema, las resuelve con un
modelo exacto de Gurobi y, si Gurobi encuentra una solución factible pero no
prueba optimalidad, intenta mejorarla mediante una metaheurística ALNS.

El problema representa la planificación de un turno de producción:
- Hay varias líneas paralelas.
- Cada línea puede abrirse o quedar cerrada.
- Una línea abierta debe alcanzar una carga mínima y no superar una carga máxima.
- Cada order tiene unas toneladas asociadas.
- Una order puede ser compatible solo con algunas líneas.
- Las orders asignadas a una misma línea deben secuenciarse.
- Algunas orders pueden quedar pendientes para el siguiente turno.

CAMBIOS PRINCIPALES DE ESTA VERSIÓN
-----------------------------------
1. Se permite definir varias configuraciones dentro de una misma ejecución.
2. Cada configuración contiene todos los parámetros modificables del experimento.
3. El Excel no guarda el código numérico interno de Gurobi.
4. El Excel guarda un estado simplificado:
      - Solución óptima
      - Solución factible
      - Sin solución factible
5. El output de texto guarda el motivo de parada con más detalle.
6. El Excel incorpora una columna resultado_final coloreada automáticamente.
7. Puede activarse o desactivarse la opción de emparejar instancias entre
   configuraciones mediante MATCH_INSTANCES_BETWEEN_CONFIGS.

ESTRUCTURA DEL CÓDIGO
---------------------
1. Configuración general del experimento.
2. Generación automática de instancias.
3. Guardado de inputs generados.
4. Funciones auxiliares de solución y evaluación.
5. Operadores destroy del ALNS.
6. Operadores repair del ALNS.
7. Construcción y resolución del modelo exacto de Gurobi.
8. Escritura del output detallado de cada instancia.
9. Escritura incremental del Excel resumen.
10. Ejecución principal.
"""

import os
import random
import math
import time

import gurobipy as gp
from gurobipy import GRB
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
try:
    from openpyxl.drawing.image import Image as ExcelImage
except Exception:
    ExcelImage = None

try:
    import matplotlib.pyplot as plt
except Exception:
    plt = None


# ============================================================
# 1. CONFIGURACIÓN GENERAL DEL EXPERIMENTO
# ============================================================
# En este bloque se definen las carpetas, la reproducibilidad y las
# configuraciones del experimento. La idea es que todos los parámetros que se
# quieran cambiar estén aquí arriba, sin tener que tocar la formulación.

CARPETA_RESULTADOS = "resultados_experimento_1_escalabilidad"
CARPETA_INPUTS = os.path.join(CARPETA_RESULTADOS, "inputs")
CARPETA_OUTPUTS = os.path.join(CARPETA_RESULTADOS, "outputs")
EXCEL_RESUMEN = os.path.join(CARPETA_RESULTADOS, "resumen_resultados.xlsx")

# Reproducibilidad general.
USE_SEED = False
SEED = 123

# Si es True, la instancia local 1 de cada configuración usa la misma semilla,
# la instancia local 2 de cada configuración usa la misma semilla, etc.
# Esto sirve para comparar configuraciones manteniendo la misma base aleatoria
# cuando tenga sentido. Si es False, cada configuración genera instancias
# independientes.
MATCH_INSTANCES_BETWEEN_CONFIGS = False

# Log de Gurobi en la terminal de VS Code.
GUROBI_OUTPUT_FLAG = 1
GUROBI_LOG_TO_CONSOLE = 1

# Mensajes de avance propios del programa.
PRINT_PROGRESS = True

# ----------------------------------------------------------------
# CONFIGURACIONES DEL EXPERIMENTO
# ----------------------------------------------------------------
# Cada diccionario representa un bloque de instancias con unos parámetros.
# Para hacer 50 instancias con 75 orders, 50 con 200 y 50 con 300, basta con
# crear tres configuraciones cambiando NUM_ORDERS y num_instancias.
#
# No se añade una columna de nombre de configuración al Excel: cada fila ya
# contiene los valores de los parámetros usados, por lo que se puede filtrar
# directamente por num_orders, num_lines, prob_coste_cero, etc.

CONFIGURACIONES = []
LISTA_ORDERS = [20, 40, 60, 80, 100, 150, 200, 250, 300, 350, 400, 450]

for n_orders in LISTA_ORDERS:
    CONFIGURACIONES.append(
        {
            "nombre": f"{n_orders}_orders",
            "num_instancias": 50,

            "NUM_LINES": 4,
            "NUM_ORDERS": n_orders,

            "MEDIA_TONELADAS": 20,
            "DESV_TONELADAS": 4,

            "MIN_LOAD_OPEN_LINE": 400,
            "MAX_LOAD_OPEN_LINE": 425,

            "PROB_COMPATIBILIDAD": 0.85,

            # 0,05% de probabilidad de coste cero.
            "PROB_COSTE_CERO": 0.0005,
            "MIN_COSTE_SETUP": 10,
            "MAX_COSTE_SETUP": 100,

            # 0,05% de probabilidad de coste inicial cero.
            "PROB_COSTE_INICIO_CERO": 0.0005,
            "MIN_COSTE_INICIO": 10,
            "MAX_COSTE_INICIO": 100,

            "PENALIZACION_LINEA_CERRADA": 1000,

            # Experimento 1: Gurobi se corta por tiempo máximo o por optimalidad.
            "MIN_TIME": 120,
            "MAX_TIME": 120,
            "TARGET_GAP": 0,

            # Se mantiene el ALNS en el código, pero se desactiva con tiempo 0.
            "ALNS_MAX_ITER": 0,
            "ALNS_MAX_TIME": 0,
            "ALNS_PORCENTAJE_REMOVAL": 0.10,
        }
    )


def total_instancias_experimento():
    """Devuelve el número total de instancias que se resolverán."""
    return sum(config["num_instancias"] for config in CONFIGURACIONES)


# ============================================================
# 2. FUNCIONES DE GENERACIÓN DE INSTANCIAS
# ============================================================

def preparar_carpetas():
    """Crea la estructura de carpetas del experimento."""
    os.makedirs(CARPETA_RESULTADOS, exist_ok=True)
    os.makedirs(CARPETA_INPUTS, exist_ok=True)
    os.makedirs(CARPETA_OUTPUTS, exist_ok=True)


def obtener_seed_instancia(indice_config, indice_local, instancia_id):
    """
    Calcula la semilla de una instancia.

    - Si USE_SEED=False, devuelve None y se usará aleatoriedad no fijada.
    - Si MATCH_INSTANCES_BETWEEN_CONFIGS=True, las instancias con el mismo
      índice local en distintas configuraciones comparten semilla.
    - Si MATCH_INSTANCES_BETWEEN_CONFIGS=False, cada instancia global recibe
      una semilla distinta.
    """
    if not USE_SEED:
        return None

    if MATCH_INSTANCES_BETWEEN_CONFIGS:
        return SEED + indice_local

    return SEED + 100000 * indice_config + instancia_id


def generar_pesos(orders, config, rng):
    """Genera las toneladas de cada order como enteros positivos."""
    weight = {}
    media = config["MEDIA_TONELADAS"]
    desv = config["DESV_TONELADAS"]

    for i in orders:
        toneladas = round(abs(rng.normalvariate(media, desv)))
        if toneladas == 0:
            toneladas = 1
        weight[i] = toneladas

    return weight


def generar_compatibilidad(orders, lines, config, rng):
    """
    Genera una matriz de compatibilidad order-línea.
    Si una order queda incompatible con todas las líneas, se fuerza
    aleatoriamente una línea compatible.
    """
    compatible = {}
    prob = config["PROB_COMPATIBILIDAD"]

    for i in orders:
        lineas_compatibles = []

        for l in lines:
            if rng.random() <= prob:
                compatible[(i, l)] = 1
                lineas_compatibles.append(l)
            else:
                compatible[(i, l)] = 0

        if len(lineas_compatibles) == 0:
            linea_forzada = rng.choice(lines)
            compatible[(i, linea_forzada)] = 1

    return compatible


def generar_matriz_costes(orders, config, rng):
    """
    Genera una matriz asimétrica de costes entre orders.

    La parte superior de la matriz se genera primero. Si el coste superior es
    cero, el coste inverso también se fija a cero. Si no es cero, el coste
    inverso se genera independientemente.
    """
    coste = {}
    prob_cero = config["PROB_COSTE_CERO"]
    min_coste = config["MIN_COSTE_SETUP"]
    max_coste = config["MAX_COSTE_SETUP"]

    for i in orders:
        for j in orders:
            if i == j:
                coste[(i, j)] = 0

    for pos_i, i in enumerate(orders):
        for j in orders[pos_i + 1:]:
            if rng.random() <= prob_cero:
                coste[(i, j)] = 0
                coste[(j, i)] = 0
            else:
                coste[(i, j)] = rng.randint(min_coste, max_coste)
                coste[(j, i)] = rng.randint(min_coste, max_coste)

    return coste


def generar_costes_inicio(orders, lines, config, rng):
    """
    Genera el coste de empezar cada línea con cada order.
    Un coste cero representa que la línea ya está preparada para esa order.
    """
    coste_inicio = {}
    prob_cero = config["PROB_COSTE_INICIO_CERO"]
    min_coste = config["MIN_COSTE_INICIO"]
    max_coste = config["MAX_COSTE_INICIO"]

    for i in orders:
        for l in lines:
            if rng.random() <= prob_cero:
                coste_inicio[(i, l)] = 0
            else:
                coste_inicio[(i, l)] = rng.randint(min_coste, max_coste)

    return coste_inicio


def generar_instancia(instancia_id, indice_config, indice_local, config):
    """Genera todos los datos necesarios para una instancia."""
    seed_instancia = obtener_seed_instancia(indice_config, indice_local, instancia_id)
    rng = random.Random(seed_instancia) if seed_instancia is not None else random.Random()

    orders = list(range(config["NUM_ORDERS"]))
    lines = list(range(config["NUM_LINES"]))

    instancia = {
        "instancia_id": instancia_id,
        "indice_config": indice_config,
        "indice_local": indice_local,
        "nombre_configuracion": config["nombre"],
        "num_orders": config["NUM_ORDERS"],
        "num_lines": config["NUM_LINES"],
        "orders": orders,
        "lines": lines,
        "weight": generar_pesos(orders, config, rng),
        "compatible": None,
        "coste": None,
        "coste_inicio": None,
        "min_load_open_line": config["MIN_LOAD_OPEN_LINE"],
        "max_load_open_line": config["MAX_LOAD_OPEN_LINE"],
        "penalizacion_linea_cerrada": config["PENALIZACION_LINEA_CERRADA"],
        "media_toneladas": config["MEDIA_TONELADAS"],
        "desv_toneladas": config["DESV_TONELADAS"],
        "prob_compatibilidad": config["PROB_COMPATIBILIDAD"],
        "prob_coste_cero": config["PROB_COSTE_CERO"],
        "min_coste_setup": config["MIN_COSTE_SETUP"],
        "max_coste_setup": config["MAX_COSTE_SETUP"],
        "prob_coste_inicio_cero": config["PROB_COSTE_INICIO_CERO"],
        "min_coste_inicio": config["MIN_COSTE_INICIO"],
        "max_coste_inicio": config["MAX_COSTE_INICIO"],
        "min_time": config["MIN_TIME"],
        "max_time": config["MAX_TIME"],
        "target_gap": config["TARGET_GAP"],
        "alns_max_iter": config["ALNS_MAX_ITER"],
        "alns_max_time": config["ALNS_MAX_TIME"],
        "alns_porcentaje_removal": config["ALNS_PORCENTAJE_REMOVAL"],
        "use_seed": USE_SEED,
        "seed_base": SEED if USE_SEED else "NO_SEED",
        "seed_instancia": seed_instancia if seed_instancia is not None else "NO_SEED",
        "match_instances_between_configs": MATCH_INSTANCES_BETWEEN_CONFIGS,
    }

    instancia["compatible"] = generar_compatibilidad(orders, lines, config, rng)
    instancia["coste"] = generar_matriz_costes(orders, config, rng)
    instancia["coste_inicio"] = generar_costes_inicio(orders, lines, config, rng)

    return instancia


# ============================================================
# 3. GUARDADO DE INPUTS GENERADOS
# ============================================================

def ruta_input(instancia_id):
    return os.path.join(CARPETA_INPUTS, f"input_{instancia_id:03d}.txt")


def ruta_output(instancia_id):
    return os.path.join(CARPETA_OUTPUTS, f"output_{instancia_id:03d}.txt")


def guardar_input_generado(instancia):
    """Guarda en texto la instancia generada para poder revisarla."""
    instancia_id = instancia["instancia_id"]
    orders = instancia["orders"]
    lines = instancia["lines"]
    weight = instancia["weight"]
    compatible = instancia["compatible"]
    coste = instancia["coste"]
    coste_inicio = instancia["coste_inicio"]

    with open(ruta_input(instancia_id), "w", encoding="utf-8") as f:
        f.write(f"# INPUT GENERADO AUTOMATICAMENTE - INSTANCIA {instancia_id}\n")
        f.write(f"# CONFIGURACION {instancia['nombre_configuracion']}\n")
        f.write(f"# INDICE_LOCAL {instancia['indice_local']}\n")
        f.write(f"# USE_SEED {instancia['use_seed']}\n")
        f.write(f"# SEED_BASE {instancia['seed_base']}\n")
        f.write(f"# SEED_INSTANCIA {instancia['seed_instancia']}\n")
        f.write(f"# MATCH_INSTANCES_BETWEEN_CONFIGS {instancia['match_instances_between_configs']}\n")
        f.write(f"# MEDIA_TONELADAS {instancia['media_toneladas']}\n")
        f.write(f"# DESV_TONELADAS {instancia['desv_toneladas']}\n")
        f.write(f"# MIN_LOAD_OPEN_LINE {instancia['min_load_open_line']}\n")
        f.write(f"# MAX_LOAD_OPEN_LINE {instancia['max_load_open_line']}\n")
        f.write(f"# PENALIZACION_LINEA_CERRADA {instancia['penalizacion_linea_cerrada']}\n")
        f.write(f"# PROB_COMPATIBILIDAD {instancia['prob_compatibilidad']}\n")
        f.write(f"# PROB_COSTE_CERO {instancia['prob_coste_cero']}\n")
        f.write(f"# PROB_COSTE_INICIO_CERO {instancia['prob_coste_inicio_cero']}\n\n")

        f.write(f"NUM_ORDERS {instancia['num_orders']}\n")
        f.write(f"NUM_LINES {instancia['num_lines']}\n\n")

        f.write("WEIGHT\n")
        for i in orders:
            f.write(f"{i} {weight[i]}\n")
        f.write("\n")

        f.write("COMPATIBILITY\n")
        for i in orders:
            fila = [str(compatible[(i, l)]) for l in lines]
            f.write(f"{i} " + " ".join(fila) + "\n")
        f.write("\n")

        f.write("COST_MATRIX\n")
        f.write("order " + " ".join(str(j) for j in orders) + "\n")
        for i in orders:
            fila = [str(coste[(i, j)]) for j in orders]
            f.write(f"{i} " + " ".join(fila) + "\n")
        f.write("\n")

        f.write("START_COST\n")
        f.write("order " + " ".join(f"line_{l}" for l in lines) + "\n")
        for i in orders:
            fila = [str(coste_inicio[(i, l)]) for l in lines]
            f.write(f"{i} " + " ".join(fila) + "\n")


# ============================================================
# 4. FUNCIONES AUXILIARES DE SOLUCIÓN
# ============================================================

I = "inicio"
F = "final"


def extraer_solucion_desde_gurobi(x, y, open_line, orders, lines):
    """
    Reconstruye las secuencias a partir de los arcos activos de Gurobi y valida
    que coinciden con las asignaciones y[i,l].

    Devuelve:
    - solucion: diccionario línea -> secuencia de orders.
    - diagnostico: información de validación por línea y validación global.

    Esta función es crítica: el modelo se basa en asignar y secuenciar orders.
    Por eso no solo reconstruye la secuencia, sino que comprueba que no haya
    orders asignadas que queden fuera de la cadena extraída.
    """
    solucion = {}
    detalle_lineas = {}
    valida_global = True
    mensajes_globales = []

    for l in lines:
        starts = [j for j in orders if x[I, j, l].x > 0.5]
        ends = [i for i in orders if x[i, F, l].x > 0.5]
        asignadas_y = [i for i in orders if y[i, l].x > 0.5]
        abierta = open_line[l].x > 0.5

        sucesores = {}
        predecesores = {}
        arcos_internos = []

        for i in orders:
            for j in orders:
                if i != j and x[i, j, l].x > 0.5:
                    arcos_internos.append((i, j))
                    sucesores.setdefault(i, []).append(j)
                    predecesores.setdefault(j, []).append(i)

        secuencia = []
        errores = []

        if not abierta:
            # Una línea cerrada no debería tener inicios, finales, arcos internos ni asignaciones.
            if starts or ends or arcos_internos or asignadas_y:
                errores.append(
                    "Línea cerrada con arcos activos o pedidos asignados."
                )
            solucion[l] = []
        else:
            if len(starts) != 1:
                errores.append(f"Número de arcos de inicio distinto de 1: {len(starts)}")
            if len(ends) != 1:
                errores.append(f"Número de arcos de final distinto de 1: {len(ends)}")

            if len(starts) == 1:
                actual = starts[0]
                visitados = set()

                while True:
                    if actual in visitados:
                        errores.append(f"Ciclo detectado al reconstruir la línea {l}.")
                        break

                    secuencia.append(actual)
                    visitados.add(actual)

                    if actual in ends:
                        break

                    sigs = sucesores.get(actual, [])

                    if len(sigs) == 0:
                        errores.append(
                            f"La reconstrucción se detiene en order {actual} sin llegar al final."
                        )
                        break

                    if len(sigs) > 1:
                        errores.append(
                            f"Order {actual} tiene más de un sucesor activo en la línea {l}: {sigs}"
                        )
                        break

                    actual = sigs[0]

            solucion[l] = secuencia

            set_sec = set(secuencia)
            set_y = set(asignadas_y)

            faltan = sorted(set_y - set_sec)
            sobran = sorted(set_sec - set_y)

            if faltan:
                errores.append(
                    f"Orders asignadas por y pero no reconstruidas en la secuencia: {faltan}"
                )
            if sobran:
                errores.append(
                    f"Orders reconstruidas en la secuencia pero no asignadas por y: {sobran}"
                )

            for pedido, preds in predecesores.items():
                if len(preds) > 1:
                    errores.append(
                        f"Order {pedido} tiene más de un predecesor activo en la línea {l}: {preds}"
                    )

        if errores:
            valida_global = False
            mensajes_globales.extend([f"Línea {l}: {e}" for e in errores])

        detalle_lineas[l] = {
            "abierta": abierta,
            "starts": starts,
            "ends": ends,
            "asignadas_y": asignadas_y,
            "secuencia": secuencia,
            "arcos_internos": arcos_internos,
            "errores": errores,
        }

    asignadas_por_y = []
    for l in lines:
        asignadas_por_y.extend(detalle_lineas[l]["asignadas_y"])

    asignadas_por_secuencia = pedidos_asignados(solucion)

    if len(asignadas_por_y) != len(set(asignadas_por_y)):
        valida_global = False
        mensajes_globales.append("Una order aparece asignada por y en más de una línea.")

    if len(asignadas_por_secuencia) != len(set(asignadas_por_secuencia)):
        valida_global = False
        mensajes_globales.append("Una order aparece repetida en las secuencias reconstruidas.")

    diagnostico = {
        "secuencias_validas": valida_global,
        "mensajes": mensajes_globales,
        "detalle_lineas": detalle_lineas,
        "num_orders_asignadas_y": len(asignadas_por_y),
        "num_orders_en_secuencias": len(asignadas_por_secuencia),
    }

    return solucion, diagnostico


def coste_setup_desde_variables_x(x, orders, lines, coste, coste_inicio):
    """
    Calcula el coste de setup directamente desde los arcos activos x de Gurobi.

    Esta métrica no depende de la reconstrucción de secuencias. Sirve como
    comprobación independiente del coste de setup obtenido a partir de la
    secuencia reconstruida.
    """
    total = 0

    for l in lines:
        for j in orders:
            if x[I, j, l].x > 0.5:
                total += coste_inicio[(j, l)]

        for i in orders:
            for j in orders:
                if i != j and x[i, j, l].x > 0.5:
                    total += coste[(i, j)]

    return total

def pedidos_asignados(solucion):
    asignados = []
    for secuencia in solucion.values():
        asignados.extend(secuencia)
    return asignados


def pedidos_pendientes(solucion, orders):
    asignados = set(pedidos_asignados(solucion))
    return [i for i in orders if i not in asignados]


def toneladas_asignadas(solucion, weight):
    return sum(weight[i] for i in pedidos_asignados(solucion))


def carga_solucion(solucion, weight):
    return {l: sum(weight[i] for i in secuencia) for l, secuencia in solucion.items()}


def coste_solucion(solucion, coste, coste_inicio):
    total = 0

    for l, secuencia in solucion.items():
        if len(secuencia) == 0:
            continue

        total += coste_inicio[(secuencia[0], l)]

        for pos in range(len(secuencia) - 1):
            i = secuencia[pos]
            j = secuencia[pos + 1]
            total += coste[(i, j)]

    return total


def numero_lineas_abiertas(solucion):
    return sum(1 for secuencia in solucion.values() if len(secuencia) > 0)


def lineas_cerradas(solucion, lines):
    return len(lines) - numero_lineas_abiertas(solucion)


def es_factible(solucion, orders, lines, weight, min_load, max_load, compatible):
    asignados = pedidos_asignados(solucion)

    if len(asignados) != len(set(asignados)):
        return False

    if any(i not in orders for i in asignados):
        return False

    cargas = carga_solucion(solucion, weight)

    for l, secuencia in solucion.items():
        carga_l = cargas[l]

        if len(secuencia) == 0:
            if abs(carga_l) > 1e-9:
                return False
        else:
            if carga_l < min_load - 1e-9:
                return False
            if carga_l > max_load + 1e-9:
                return False

        for i in secuencia:
            if compatible[(i, l)] == 0:
                return False

    return True


def objetivo_total(solucion, orders, lines, coste, coste_inicio, penalizacion_linea_cerrada):
    return (
        coste_solucion(solucion, coste, coste_inicio)
        + penalizacion_linea_cerrada * lineas_cerradas(solucion, lines)
    )


def elegir_por_pesos(pesos):
    total = sum(pesos.values())

    if total <= 0:
        return random.choice(list(pesos.keys()))

    r = random.uniform(0, total)
    acumulado = 0

    for nombre, peso in pesos.items():
        acumulado += peso
        if r <= acumulado:
            return nombre

    return list(pesos.keys())[-1]


# ============================================================
# 5. ALNS: OPERADORES DESTROY
# ============================================================

def worst_removal(solucion, q, coste, coste_inicio):
    coste_base = coste_solucion(solucion, coste, coste_inicio)
    contribuciones = []

    for l, secuencia in solucion.items():
        for pedido in secuencia:
            candidata = {ll: ss.copy() for ll, ss in solucion.items()}
            candidata[l].remove(pedido)
            nuevo_coste = coste_solucion(candidata, coste, coste_inicio)
            ahorro = coste_base - nuevo_coste
            contribuciones.append((ahorro, l, pedido))

    contribuciones.sort(reverse=True)

    nueva_solucion = {l: sec.copy() for l, sec in solucion.items()}
    eliminados = []

    for _, l, pedido in contribuciones[:q]:
        if pedido in nueva_solucion[l]:
            nueva_solucion[l].remove(pedido)
            eliminados.append(pedido)

    return nueva_solucion, eliminados


def random_removal(solucion, q, coste, coste_inicio):
    todos = []

    for l, secuencia in solucion.items():
        for pedido in secuencia:
            todos.append((l, pedido))

    q_real = min(q, len(todos))

    if q_real == 0:
        return {l: sec.copy() for l, sec in solucion.items()}, []

    seleccionados = random.sample(todos, q_real)

    nueva_solucion = {l: sec.copy() for l, sec in solucion.items()}
    eliminados = []

    for l, pedido in seleccionados:
        if pedido in nueva_solucion[l]:
            nueva_solucion[l].remove(pedido)
            eliminados.append(pedido)

    return nueva_solucion, eliminados


def related_removal(solucion, q, coste, coste_inicio):
    """
    Elimina orders relacionadas por bajo coste de transición.
    Como ya no existen setup_in/setup_out, se usa directamente la matriz de costes.
    """
    todos = []

    for secuencia in solucion.values():
        todos.extend(secuencia)

    if len(todos) == 0:
        return {l: sec.copy() for l, sec in solucion.items()}, []

    q_real = min(q, len(todos))
    semilla = random.choice(todos)
    eliminados_set = {semilla}

    while len(eliminados_set) < q_real:
        referencia = random.choice(list(eliminados_set))
        candidatos = [p for p in todos if p not in eliminados_set]

        if not candidatos:
            break

        candidatos.sort(key=lambda p: min(coste[(referencia, p)], coste[(p, referencia)]))
        limite = max(1, min(5, len(candidatos)))
        elegido = random.choice(candidatos[:limite])
        eliminados_set.add(elegido)

    nueva_solucion = {l: sec.copy() for l, sec in solucion.items()}
    eliminados = []

    for pedido in eliminados_set:
        for l in nueva_solucion.keys():
            if pedido in nueva_solucion[l]:
                nueva_solucion[l].remove(pedido)
                eliminados.append(pedido)
                break

    return nueva_solucion, eliminados


def close_line_removal(solucion, q, coste, coste_inicio):
    """Cierra una línea completa para explorar soluciones con menos o distintas líneas abiertas."""
    abiertas = [l for l, sec in solucion.items() if len(sec) > 0]

    if not abiertas:
        return {l: sec.copy() for l, sec in solucion.items()}, []

    l_cerrar = random.choice(abiertas)
    nueva_solucion = {l: sec.copy() for l, sec in solucion.items()}
    eliminados = nueva_solucion[l_cerrar].copy()
    nueva_solucion[l_cerrar] = []

    return nueva_solucion, eliminados


# ============================================================
# 6. ALNS: OPERADORES REPAIR
# ============================================================

def mejor_insercion_pedido(solucion, pedido, orders, lines, coste, coste_inicio,
                           weight, min_load, max_load, compatible,
                           penalizacion_linea_cerrada):
    opciones = []

    if pedido in pedidos_asignados(solucion):
        return []

    for l in lines:
        if compatible[(pedido, l)] == 0:
            continue

        carga_actual = sum(weight[i] for i in solucion[l])

        if carga_actual + weight[pedido] > max_load + 1e-9:
            continue

        for pos in range(len(solucion[l]) + 1):
            candidata = {ll: ss.copy() for ll, ss in solucion.items()}
            candidata[l].insert(pos, pedido)

            obj = objetivo_total(
                candidata,
                orders,
                lines,
                coste,
                coste_inicio,
                penalizacion_linea_cerrada
            )

            opciones.append((obj, l, pos))

    opciones.sort(key=lambda x: x[0])
    return opciones


def cheapest_insertion(solucion_parcial, candidatos_insertar, orders, lines, coste, coste_inicio,
                       weight, min_load, max_load, compatible, penalizacion_linea_cerrada):
    solucion = {l: sec.copy() for l, sec in solucion_parcial.items()}

    for pedido in candidatos_insertar:
        if pedido in pedidos_asignados(solucion):
            continue

        opciones = mejor_insercion_pedido(
            solucion,
            pedido,
            orders,
            lines,
            coste,
            coste_inicio,
            weight,
            min_load,
            max_load,
            compatible,
            penalizacion_linea_cerrada
        )

        if len(opciones) == 0:
            continue

        _, mejor_linea, mejor_pos = opciones[0]
        solucion[mejor_linea].insert(mejor_pos, pedido)

    return solucion


def regret_2_insertion(solucion_parcial, candidatos_insertar, orders, lines, coste, coste_inicio,
                       weight, min_load, max_load, compatible, penalizacion_linea_cerrada):
    solucion = {l: sec.copy() for l, sec in solucion_parcial.items()}

    pendientes = list(dict.fromkeys(
        p for p in candidatos_insertar
        if p not in pedidos_asignados(solucion)
    ))

    while pendientes:
        mejor_pedido = None
        mejor_linea = None
        mejor_pos = None
        mejor_regret = -float("inf")
        mejor_obj = float("inf")

        for pedido in pendientes:
            opciones = mejor_insercion_pedido(
                solucion,
                pedido,
                orders,
                lines,
                coste,
                coste_inicio,
                weight,
                min_load,
                max_load,
                compatible,
                penalizacion_linea_cerrada
            )

            if len(opciones) == 0:
                continue

            obj_1, linea_1, pos_1 = opciones[0]

            if len(opciones) >= 2:
                obj_2 = opciones[1][0]
                regret = obj_2 - obj_1
            else:
                regret = float("inf")

            if regret > mejor_regret or (regret == mejor_regret and obj_1 < mejor_obj):
                mejor_regret = regret
                mejor_obj = obj_1
                mejor_pedido = pedido
                mejor_linea = linea_1
                mejor_pos = pos_1

        if mejor_pedido is None:
            break

        solucion[mejor_linea].insert(mejor_pos, mejor_pedido)
        pendientes.remove(mejor_pedido)

    return solucion


def reparar_lineas_incompletas(solucion, orders, lines, coste, coste_inicio, weight,
                               min_load, max_load, compatible, penalizacion_linea_cerrada):
    """
    Repara líneas abiertas con carga positiva pero inferior al mínimo.
    Si no puede completarlas, la solución se descartará como no factible.
    """
    sol = {l: sec.copy() for l, sec in solucion.items()}
    pendientes = pedidos_pendientes(sol, orders)

    cambio = True
    while cambio:
        cambio = False
        cargas = carga_solucion(sol, weight)

        lineas_incompletas = [
            l for l in lines
            if 0 < cargas[l] < min_load - 1e-9
        ]

        if not lineas_incompletas:
            break

        for l in lineas_incompletas:
            mejor_pedido = None
            mejor_pos = None
            mejor_obj = float("inf")

            for pedido in pendientes:
                if compatible[(pedido, l)] == 0:
                    continue

                if cargas[l] + weight[pedido] > max_load + 1e-9:
                    continue

                for pos in range(len(sol[l]) + 1):
                    candidata = {ll: ss.copy() for ll, ss in sol.items()}
                    candidata[l].insert(pos, pedido)

                    obj = objetivo_total(
                        candidata,
                        orders,
                        lines,
                        coste,
                        coste_inicio,
                        penalizacion_linea_cerrada
                    )

                    if obj < mejor_obj:
                        mejor_obj = obj
                        mejor_pedido = pedido
                        mejor_pos = pos

            if mejor_pedido is not None:
                sol[l].insert(mejor_pos, mejor_pedido)
                pendientes.remove(mejor_pedido)
                cambio = True

    return sol


def alns(solucion_inicial, orders, lines, coste, coste_inicio, weight,
         min_load, max_load, compatible, penalizacion_linea_cerrada,
         max_iter=500000, max_time=300, porcentaje_removal=0.20):
    tiempo_inicio = time.time()

    actual = {l: sec.copy() for l, sec in solucion_inicial.items()}
    mejor = {l: sec.copy() for l, sec in solucion_inicial.items()}

    obj_actual = objetivo_total(actual, orders, lines, coste, coste_inicio, penalizacion_linea_cerrada)
    mejor_obj = obj_actual

    q = max(1, int(len(orders) * porcentaje_removal))

    temperatura = max(1, 0.05 * obj_actual)
    enfriamiento = 0.995

    removal_operators = {
        "worst_removal": worst_removal,
        "random_removal": random_removal,
        "related_removal": related_removal,
        "close_line_removal": close_line_removal
    }

    repair_operators = {
        "cheapest_insertion": cheapest_insertion,
        "regret_2_insertion": regret_2_insertion
    }

    pesos_removal = {nombre: 1.0 for nombre in removal_operators}
    pesos_repair = {nombre: 1.0 for nombre in repair_operators}

    puntuacion_removal = {nombre: 0.0 for nombre in removal_operators}
    puntuacion_repair = {nombre: 0.0 for nombre in repair_operators}

    usos_removal = {nombre: 0 for nombre in removal_operators}
    usos_repair = {nombre: 0 for nombre in repair_operators}

    segmento = 10
    factor_reaccion = 0.20
    premio_mejor_global = 10
    premio_mejora_actual = 5
    premio_aceptada_peor = 1

    iteraciones_realizadas = 0

    for _ in range(max_iter):
        if time.time() - tiempo_inicio >= max_time:
            break

        iteraciones_realizadas += 1

        nombre_removal = elegir_por_pesos(pesos_removal)
        nombre_repair = elegir_por_pesos(pesos_repair)

        usos_removal[nombre_removal] += 1
        usos_repair[nombre_repair] += 1

        parcial, eliminados = removal_operators[nombre_removal](actual, q, coste, coste_inicio)

        pendientes = pedidos_pendientes(parcial, orders)
        random.shuffle(pendientes)
        candidatos_extra = pendientes[:q]
        candidatos_insertar = list(dict.fromkeys(eliminados + candidatos_extra))

        candidata = repair_operators[nombre_repair](
            parcial,
            candidatos_insertar,
            orders,
            lines,
            coste,
            coste_inicio,
            weight,
            min_load,
            max_load,
            compatible,
            penalizacion_linea_cerrada
        )

        candidata = reparar_lineas_incompletas(
            candidata,
            orders,
            lines,
            coste,
            coste_inicio,
            weight,
            min_load,
            max_load,
            compatible,
            penalizacion_linea_cerrada
        )

        if not es_factible(candidata, orders, lines, weight, min_load, max_load, compatible):
            continue

        obj_candidata = objetivo_total(candidata, orders, lines, coste, coste_inicio, penalizacion_linea_cerrada)
        diferencia = obj_candidata - obj_actual

        aceptar = False
        premio = 0

        if obj_candidata < mejor_obj:
            aceptar = True
            premio = premio_mejor_global
        elif diferencia < 0:
            aceptar = True
            premio = premio_mejora_actual
        else:
            probabilidad = math.exp(-diferencia / temperatura) if temperatura > 1e-9 else 0
            aceptar = random.random() < probabilidad

            if aceptar:
                premio = premio_aceptada_peor

        if premio > 0:
            puntuacion_removal[nombre_removal] += premio
            puntuacion_repair[nombre_repair] += premio

        if aceptar:
            actual = candidata
            obj_actual = obj_candidata

        if obj_candidata < mejor_obj:
            mejor = candidata
            mejor_obj = obj_candidata

        temperatura *= enfriamiento

        if iteraciones_realizadas % segmento == 0:
            for nombre in pesos_removal:
                if usos_removal[nombre] > 0:
                    rendimiento = puntuacion_removal[nombre] / usos_removal[nombre]
                    pesos_removal[nombre] = (
                        (1 - factor_reaccion) * pesos_removal[nombre]
                        + factor_reaccion * rendimiento
                    )
                puntuacion_removal[nombre] = 0.0
                usos_removal[nombre] = 0

            for nombre in pesos_repair:
                if usos_repair[nombre] > 0:
                    rendimiento = puntuacion_repair[nombre] / usos_repair[nombre]
                    pesos_repair[nombre] = (
                        (1 - factor_reaccion) * pesos_repair[nombre]
                        + factor_reaccion * rendimiento
                    )
                puntuacion_repair[nombre] = 0.0
                usos_repair[nombre] = 0

            for nombre in pesos_removal:
                pesos_removal[nombre] = max(0.05, pesos_removal[nombre])

            for nombre in pesos_repair:
                pesos_repair[nombre] = max(0.05, pesos_repair[nombre])

    tiempo_total = time.time() - tiempo_inicio

    estadisticas = {
        "iteraciones": iteraciones_realizadas,
        "tiempo": tiempo_total,
        "objetivo_inicial": objetivo_total(solucion_inicial, orders, lines, coste, coste_inicio, penalizacion_linea_cerrada),
        "mejor_objetivo": mejor_obj,
        "coste_setup_inicial": coste_solucion(solucion_inicial, coste, coste_inicio),
        "coste_setup_mejor": coste_solucion(mejor, coste, coste_inicio),
        "toneladas_iniciales": toneladas_asignadas(solucion_inicial, weight),
        "toneladas_mejor": toneladas_asignadas(mejor, weight),
        "lineas_abiertas_inicial": numero_lineas_abiertas(solucion_inicial),
        "lineas_abiertas_mejor": numero_lineas_abiertas(mejor),
        "pesos_removal": pesos_removal,
        "pesos_repair": pesos_repair
    }

    return mejor, mejor_obj, estadisticas


# ============================================================
# 7. MODELO EXACTO GUROBI
# ============================================================

def resolver_instancia(instancia, total_instancias):
    """
    Construye y resuelve el modelo exacto de Gurobi para una instancia.
    Si Gurobi no prueba optimalidad pero tiene solución factible, ejecuta ALNS.
    """
    instancia_id = instancia["instancia_id"]
    orders = instancia["orders"]
    lines = instancia["lines"]
    weight = instancia["weight"]
    compatible = instancia["compatible"]
    coste = instancia["coste"]
    coste_inicio = instancia["coste_inicio"]
    min_load = instancia["min_load_open_line"]
    max_load = instancia["max_load_open_line"]
    penalizacion_linea_cerrada = instancia["penalizacion_linea_cerrada"]

    m = gp.Model(f"modelo_industrial_instancia_{instancia_id:03d}")

    m.setParam("OutputFlag", GUROBI_OUTPUT_FLAG)
    m.setParam("LogToConsole", GUROBI_LOG_TO_CONSOLE)

    m._parada_por_gap_y_tiempo = False
    m._parada_por_tiempo_maximo = False

    def parada_personalizada(model, where):
        if where == GRB.Callback.MIP:
            runtime = model.cbGet(GRB.Callback.RUNTIME)

            if runtime >= instancia["max_time"]:
                model._parada_por_tiempo_maximo = True
                model.terminate()
                return

            obj_best = model.cbGet(GRB.Callback.MIP_OBJBST)
            obj_bound = model.cbGet(GRB.Callback.MIP_OBJBND)

            if obj_best < GRB.INFINITY and abs(obj_best) > 1e-9:
                gap = abs(obj_best - obj_bound) / abs(obj_best)

                if runtime >= instancia["min_time"] and gap <= instancia["target_gap"]:
                    model._parada_por_gap_y_tiempo = True
                    model.terminate()

    arcos_orders = [(i, j, l) for i in orders for j in orders if i != j for l in lines]
    arcos_inicio = [(I, j, l) for j in orders for l in lines]
    arcos_final = [(i, F, l) for i in orders for l in lines]
    arcos = arcos_orders + arcos_inicio + arcos_final

    x = m.addVars(arcos, vtype=GRB.BINARY, name="x")
    u = m.addVars(orders, lines, vtype=GRB.CONTINUOUS, lb=0, ub=len(orders), name="u")
    y = m.addVars(orders, lines, vtype=GRB.BINARY, name="assign_line")
    assigned = m.addVars(orders, vtype=GRB.BINARY, name="assigned_turn")
    open_line = m.addVars(lines, vtype=GRB.BINARY, name="open_line")
    carga = m.addVars(lines, vtype=GRB.CONTINUOUS, lb=0, name="load")

    coste_setup_total = (
        gp.quicksum(coste[(i, j)] * x[i, j, l] for i in orders for j in orders if i != j for l in lines)
        + gp.quicksum(coste_inicio[(j, l)] * x[I, j, l] for j in orders for l in lines)
    )

    lineas_cerradas_modelo = gp.quicksum(1 - open_line[l] for l in lines)

    m.setObjective(
        coste_setup_total + penalizacion_linea_cerrada * lineas_cerradas_modelo,
        GRB.MINIMIZE
    )

    for i in orders:
        m.addConstr(
            gp.quicksum(x[i, j, l] for j in orders if i != j for l in lines)
            + gp.quicksum(x[i, F, l] for l in lines)
            == assigned[i],
            name=f"salida_si_asignado_{i}"
        )

    for j in orders:
        m.addConstr(
            gp.quicksum(x[i, j, l] for i in orders if i != j for l in lines)
            + gp.quicksum(x[I, j, l] for l in lines)
            == assigned[j],
            name=f"entrada_si_asignado_{j}"
        )

    for l in lines:
        for k in orders:
            m.addConstr(
                gp.quicksum(x[i, k, l] for i in orders if i != k) + x[I, k, l]
                ==
                gp.quicksum(x[k, j, l] for j in orders if j != k) + x[k, F, l],
                name=f"flujo_linea_{l}_order_{k}"
            )

    for l in lines:
        m.addConstr(
            gp.quicksum(x[I, j, l] for j in orders) == open_line[l],
            name=f"inicio_si_linea_abierta_{l}"
        )

        m.addConstr(
            gp.quicksum(x[i, F, l] for i in orders) == open_line[l],
            name=f"final_si_linea_abierta_{l}"
        )

    for l in lines:
        for i in orders:
            m.addConstr(
                y[i, l] == gp.quicksum(x[h, i, l] for h in orders if h != i) + x[I, i, l],
                name=f"def_assign_line_{i}_{l}"
            )

    for i in orders:
        m.addConstr(
            gp.quicksum(y[i, l] for l in lines) == assigned[i],
            name=f"asignacion_total_{i}"
        )

    for i in orders:
        for l in lines:
            m.addConstr(
                y[i, l] <= compatible[(i, l)],
                name=f"compatibilidad_order_{i}_linea_{l}"
            )

    for l in lines:
        m.addConstr(
            carga[l] == gp.quicksum(weight[i] * y[i, l] for i in orders),
            name=f"def_load_{l}"
        )

        m.addConstr(
            carga[l] >= min_load * open_line[l],
            name=f"minimo_si_linea_abierta_{l}"
        )

        m.addConstr(
            carga[l] <= max_load * open_line[l],
            name=f"maximo_si_linea_abierta_{l}"
        )

    n = len(orders)
    for i in orders:
        for j in orders:
            if i != j:
                for l in lines:
                    m.addConstr(
                        u[i, l] - u[j, l] + n * x[i, j, l] <= n - 1,
                        name=f"mtz_{i}_{j}_{l}"
                    )

    if PRINT_PROGRESS:
        print("\n==================================================")
        print(f"RESOLVIENDO INSTANCIA {instancia_id}/{total_instancias}")
        print(f"Configuración: {instancia['nombre_configuracion']} | local {instancia['indice_local']}")
        print("==================================================\n")

    m.optimize(parada_personalizada)

    resultado = {
        "instancia_id": instancia_id,
        "status_gurobi": m.status,
        "solucion_gurobi": None,
        "solucion_alns": None,
        "alns_ejecutado": False,
        "stats_alns": None,
        "mip_gap": None,
        "runtime_gurobi": getattr(m, "Runtime", None),
        "obj_gurobi": None,
        "obj_alns": None,
        "parada_por_gap_y_tiempo": m._parada_por_gap_y_tiempo,
        "parada_por_tiempo_maximo": m._parada_por_tiempo_maximo,
        "sol_count": m.SolCount,
        "diagnostico_secuencias_gurobi": None,
        "secuencias_validas_gurobi": None,
        "coste_setup_gurobi_desde_x": None,
        "coste_setup_gurobi_desde_secuencia": None,
        "diferencia_coste_setup_x_vs_secuencia": None,
    }

    if m.SolCount > 0:
        solucion_gurobi, diagnostico_gurobi = extraer_solucion_desde_gurobi(
            x, y, open_line, orders, lines
        )
        coste_setup_x = coste_setup_desde_variables_x(x, orders, lines, coste, coste_inicio)
        coste_setup_secuencia = coste_solucion(solucion_gurobi, coste, coste_inicio)

        resultado["solucion_gurobi"] = solucion_gurobi
        resultado["diagnostico_secuencias_gurobi"] = diagnostico_gurobi
        resultado["secuencias_validas_gurobi"] = diagnostico_gurobi["secuencias_validas"]
        resultado["coste_setup_gurobi_desde_x"] = coste_setup_x
        resultado["coste_setup_gurobi_desde_secuencia"] = coste_setup_secuencia
        resultado["diferencia_coste_setup_x_vs_secuencia"] = coste_setup_x - coste_setup_secuencia
        resultado["obj_gurobi"] = m.objVal

        try:
            resultado["mip_gap"] = m.MIPGap
        except Exception:
            resultado["mip_gap"] = None

        if m.status != GRB.OPTIMAL and diagnostico_gurobi["secuencias_validas"]:
            if USE_SEED and instancia["seed_instancia"] != "NO_SEED":
                random.seed(instancia["seed_instancia"] + 999999)

            solucion_alns, obj_alns, stats_alns = alns(
                solucion_gurobi,
                orders,
                lines,
                coste,
                coste_inicio,
                weight,
                min_load,
                max_load,
                compatible,
                penalizacion_linea_cerrada,
                max_iter=instancia["alns_max_iter"],
                max_time=instancia["alns_max_time"],
                porcentaje_removal=instancia["alns_porcentaje_removal"]
            )
            resultado["solucion_alns"] = solucion_alns
            resultado["obj_alns"] = obj_alns
            resultado["alns_ejecutado"] = True
            resultado["stats_alns"] = stats_alns

    return resultado


# ============================================================
# 8. OUTPUT DE CADA INSTANCIA
# ============================================================

def estado_gurobi_excel(resultado):
    """Estado simplificado para el Excel."""
    if resultado["sol_count"] == 0:
        return "Sin solución factible"
    if resultado["status_gurobi"] == GRB.OPTIMAL:
        return "Solución óptima"
    return "Solución factible"


def motivo_parada_output(resultado):
    """Motivo de parada detallado para el archivo output."""
    if resultado["sol_count"] == 0:
        return "Sin solución factible"
    if resultado["status_gurobi"] == GRB.OPTIMAL:
        return "Parada por solución óptima"
    if resultado["parada_por_tiempo_maximo"]:
        return "Parada por tiempo máximo alcanzado"
    if resultado["parada_por_gap_y_tiempo"]:
        return "Parada por gap objetivo y tiempo mínimo alcanzados"
    return "Parada sin prueba de optimalidad"


def resultado_final_excel(resultado):
    """Clasificación final de la instancia para el Excel."""
    if resultado["sol_count"] == 0:
        return "Sin solución factible"

    if resultado["status_gurobi"] == GRB.OPTIMAL:
        return "Solución óptima"

    if resultado["alns_ejecutado"] and resultado["stats_alns"] is not None:
        mejora = resultado["stats_alns"]["objetivo_inicial"] - resultado["stats_alns"]["mejor_objetivo"]
        if mejora > 1e-9:
            return "Solución factible con mejora de ALNS"
        return "Solución factible sin mejora de ALNS"

    return "Sin solución óptima"


def escribir(linea, f):
    print(linea)
    f.write(linea + "\n")


def escribir_resumen_solucion(nombre, solucion, instancia, f):
    orders = instancia["orders"]
    lines = instancia["lines"]
    weight = instancia["weight"]
    coste = instancia["coste"]
    coste_inicio = instancia["coste_inicio"]
    min_load = instancia["min_load_open_line"]
    max_load = instancia["max_load_open_line"]
    compatible = instancia["compatible"]

    pendientes = pedidos_pendientes(solucion, orders)
    cargas = carga_solucion(solucion, weight)
    toneladas_totales = sum(weight[i] for i in orders)
    toneladas_proc = toneladas_asignadas(solucion, weight)
    coste_setup = coste_solucion(solucion, coste, coste_inicio)
    factible = es_factible(solucion, orders, lines, weight, min_load, max_load, compatible)

    escribir(f"Resumen {nombre}:", f)
    escribir(f"Factible: {'Sí' if factible else 'No'}", f)
    escribir(f"Toneladas totales disponibles: {toneladas_totales}", f)
    escribir(f"Toneladas asignadas al turno: {toneladas_proc}", f)
    escribir(f"Toneladas pendientes: {toneladas_totales - toneladas_proc}", f)
    escribir(f"Porcentaje toneladas procesadas: {100 * toneladas_proc / toneladas_totales if toneladas_totales > 0 else 0:.2f} %", f)
    escribir(f"Número de líneas abiertas: {numero_lineas_abiertas(solucion)}", f)
    escribir(f"Número de líneas cerradas: {lineas_cerradas(solucion, lines)}", f)
    escribir(f"Número de orders asignadas: {len(pedidos_asignados(solucion))}", f)
    escribir(f"Número de orders pendientes: {len(pendientes)}", f)
    escribir(f"Coste setup: {coste_setup}", f)
    escribir(f"Objetivo total calculado: {objetivo_total(solucion, orders, lines, coste, coste_inicio, instancia['penalizacion_linea_cerrada'])}", f)
    escribir("", f)

    escribir(f"Carga por línea según {nombre}:", f)
    for l in lines:
        estado = "abierta" if len(solucion[l]) > 0 else "cerrada"
        escribir(f"Línea {l} ({estado}): {cargas[l]} toneladas / mínimo {min_load} / máximo {max_load}", f)

    escribir("", f)
    escribir(f"Secuencias según {nombre}:", f)
    for l in lines:
        escribir(f"\nLínea {l}:", f)
        secuencia = solucion[l]
        if len(secuencia) == 0:
            escribir("  Línea cerrada", f)
        else:
            escribir("  start -> " + " -> ".join(map(str, secuencia)) + " -> end", f)

    escribir("", f)
    escribir(f"Orders pendientes por {nombre}:", f)
    escribir(", ".join(map(str, pendientes)) if pendientes else "Ninguna", f)
    escribir("", f)


def guardar_output_instancia(instancia, resultado):
    instancia_id = instancia["instancia_id"]

    with open(ruta_output(instancia_id), "w", encoding="utf-8") as f:
        escribir("==================================================", f)
        escribir("MODELO INDUSTRIAL EXPERIMENTAL", f)
        escribir("==================================================", f)
        escribir(f"Instancia: {instancia_id}", f)
        escribir(f"Configuración: {instancia['nombre_configuracion']}", f)
        escribir(f"Índice local dentro de la configuración: {instancia['indice_local']}", f)
        escribir(f"Input generado: {ruta_input(instancia_id)}", f)
        escribir(f"Output generado: {ruta_output(instancia_id)}", f)
        escribir(f"Excel resumen: {EXCEL_RESUMEN}", f)
        escribir("", f)

        escribir("Parámetros de la instancia:", f)
        escribir(f"Número de líneas: {instancia['num_lines']}", f)
        escribir(f"Número de orders: {instancia['num_orders']}", f)
        escribir(f"Carga mínima por línea abierta: {instancia['min_load_open_line']}", f)
        escribir(f"Carga máxima por línea abierta: {instancia['max_load_open_line']}", f)
        escribir(f"Penalización por línea cerrada: {instancia['penalizacion_linea_cerrada']}", f)
        escribir(f"Probabilidad compatibilidad: {instancia['prob_compatibilidad']}", f)
        escribir(f"Probabilidad coste cero entre orders: {instancia['prob_coste_cero']}", f)
        escribir(f"Probabilidad coste inicio cero: {instancia['prob_coste_inicio_cero']}", f)
        escribir(f"USE_SEED: {instancia['use_seed']}", f)
        escribir(f"SEED_BASE: {instancia['seed_base']}", f)
        escribir(f"SEED_INSTANCIA: {instancia['seed_instancia']}", f)
        escribir(f"MATCH_INSTANCES_BETWEEN_CONFIGS: {instancia['match_instances_between_configs']}", f)
        escribir("", f)

        escribir("Resultado Gurobi:", f)
        escribir(f"Estado simplificado: {estado_gurobi_excel(resultado)}", f)
        escribir(f"Motivo de parada: {motivo_parada_output(resultado)}", f)
        escribir(f"Tiempo Gurobi: {resultado['runtime_gurobi']} segundos", f)
        escribir(f"Valor objetivo Gurobi: {resultado['obj_gurobi']}", f)
        escribir(f"Gap final: {resultado['mip_gap'] if resultado['mip_gap'] is not None else 'No disponible'}", f)
        escribir(f"Resultado final: {resultado_final_excel(resultado)}", f)
        escribir("", f)

        if resultado["solucion_gurobi"] is not None:
            escribir("Validación de secuencias de Gurobi:", f)
            if resultado["secuencias_validas_gurobi"]:
                escribir("Secuencias reconstruidas correctamente: Sí", f)
            else:
                escribir("Secuencias reconstruidas correctamente: No", f)
                diagnostico = resultado["diagnostico_secuencias_gurobi"] or {}
                for mensaje in diagnostico.get("mensajes", []):
                    escribir(f"  - {mensaje}", f)
            escribir(f"Orders asignadas según y: {resultado['diagnostico_secuencias_gurobi']['num_orders_asignadas_y']}", f)
            escribir(f"Orders en secuencias reconstruidas: {resultado['diagnostico_secuencias_gurobi']['num_orders_en_secuencias']}", f)
            escribir(f"Coste setup calculado desde arcos x: {resultado['coste_setup_gurobi_desde_x']}", f)
            escribir(f"Coste setup calculado desde secuencia: {resultado['coste_setup_gurobi_desde_secuencia']}", f)
            escribir(f"Diferencia x - secuencia: {resultado['diferencia_coste_setup_x_vs_secuencia']}", f)
            escribir("", f)

            escribir_resumen_solucion("Gurobi", resultado["solucion_gurobi"], instancia, f)

            if resultado["alns_ejecutado"]:
                stats = resultado["stats_alns"]

                escribir("==================================================", f)
                escribir("MEJORA MEDIANTE ALNS", f)
                escribir("==================================================", f)
                escribir(f"Iteraciones ALNS realizadas: {stats['iteraciones']}", f)
                escribir(f"Tiempo ALNS: {stats['tiempo']} segundos", f)
                escribir(f"Objetivo inicial tomado de Gurobi: {stats['objetivo_inicial']}", f)
                escribir(f"Mejor objetivo encontrado por ALNS: {stats['mejor_objetivo']}", f)
                escribir(f"Coste setup inicial Gurobi: {stats['coste_setup_inicial']}", f)
                escribir(f"Coste setup mejor ALNS: {stats['coste_setup_mejor']}", f)
                escribir(f"Toneladas iniciales Gurobi: {stats['toneladas_iniciales']}", f)
                escribir(f"Toneladas mejor ALNS: {stats['toneladas_mejor']}", f)
                escribir(f"Líneas abiertas inicial Gurobi: {stats['lineas_abiertas_inicial']}", f)
                escribir(f"Líneas abiertas mejor ALNS: {stats['lineas_abiertas_mejor']}", f)

                mejora_obj = stats["objetivo_inicial"] - stats["mejor_objetivo"]
                porcentaje_mejora_obj = 100 * mejora_obj / stats["objetivo_inicial"] if abs(stats["objetivo_inicial"]) > 1e-9 else 0
                mejora_coste = stats["coste_setup_inicial"] - stats["coste_setup_mejor"]
                porcentaje_mejora_coste = 100 * mejora_coste / stats["coste_setup_inicial"] if abs(stats["coste_setup_inicial"]) > 1e-9 else 0

                escribir(f"Mejora absoluta objetivo ALNS: {mejora_obj}", f)
                escribir(f"Mejora porcentual objetivo ALNS: {porcentaje_mejora_obj} %", f)
                escribir(f"Mejora absoluta coste setup: {mejora_coste}", f)
                escribir(f"Mejora porcentual coste setup: {porcentaje_mejora_coste} %", f)
                escribir("", f)

                escribir_resumen_solucion("ALNS", resultado["solucion_alns"], instancia, f)
            else:
                escribir("==================================================", f)
                escribir("ALNS NO EJECUTADO", f)
                escribir("==================================================", f)
                if resultado["status_gurobi"] == GRB.OPTIMAL:
                    escribir("Gurobi ha demostrado optimalidad, por lo que no se aplica la metaheurística.", f)
                elif not resultado.get("secuencias_validas_gurobi", True):
                    escribir("ALNS no ejecutado porque la secuencia inicial de Gurobi no se ha podido validar.", f)
                else:
                    escribir("No existe solución factible inicial sobre la que aplicar ALNS.", f)
        else:
            escribir("No hay solución que resumir.", f)


# ============================================================
# 9. EXCEL RESUMEN
# ============================================================

COLUMNAS_EXCEL = [
    "instancia_id",
    "indice_local",
    "num_orders",
    "num_lines",
    "toneladas_totales",
    "media_toneladas",
    "desv_toneladas",
    "min_load_open_line",
    "max_load_open_line",
    "penalizacion_linea_cerrada",
    "prob_compatibilidad",
    "prob_coste_cero",
    "prob_coste_inicio_cero",
    "use_seed",
    "seed_base",
    "seed_instancia",
    "match_instances_between_configs",
    "estado_gurobi",
    "obj_gurobi",
    "gap_gurobi",
    "tiempo_gurobi",
    "lineas_abiertas_gurobi",
    "lineas_cerradas_gurobi",
    "toneladas_asignadas_gurobi",
    "toneladas_pendientes_gurobi",
    "porcentaje_toneladas_gurobi",
    "orders_asignadas_gurobi",
    "orders_pendientes_gurobi",
    "coste_setup_gurobi",
    "coste_setup_gurobi_desde_x",
    "coste_setup_gurobi_desde_secuencia",
    "diferencia_coste_setup_x_vs_secuencia",
    "secuencias_validas_gurobi",
    "alns_ejecutado",
    "obj_alns",
    "tiempo_alns",
    "iteraciones_alns",
    "lineas_abiertas_alns",
    "lineas_cerradas_alns",
    "toneladas_asignadas_alns",
    "toneladas_pendientes_alns",
    "porcentaje_toneladas_alns",
    "orders_asignadas_alns",
    "orders_pendientes_alns",
    "coste_setup_alns",
    "mejora_objetivo_alns",
    "mejora_porcentual_objetivo_alns",
    "mejora_coste_setup_alns",
    "mejora_porcentual_coste_setup_alns",
    "resultado_final",
    "input_file",
    "output_file"
]

COLORES_RESULTADO_FINAL = {
    "Solución óptima": "D9EAD3",
    "Solución factible con mejora de ALNS": "D9EAF7",
    "Solución factible sin mejora de ALNS": "FFF2CC",
    "Sin solución factible": "F4CCCC",
    "Sin solución óptima": "EADCF8",
}


def inicializar_excel():
    """Crea el Excel resumen si no existe."""
    if os.path.exists(EXCEL_RESUMEN):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append(COLUMNAS_EXCEL)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(EXCEL_RESUMEN)


def construir_fila_excel(instancia, resultado):
    orders = instancia["orders"]
    lines = instancia["lines"]
    weight = instancia["weight"]
    coste = instancia["coste"]
    coste_inicio = instancia["coste_inicio"]
    toneladas_totales = sum(weight[i] for i in orders)

    fila = {col: None for col in COLUMNAS_EXCEL}

    fila["instancia_id"] = instancia["instancia_id"]
    fila["indice_local"] = instancia["indice_local"]
    fila["num_orders"] = instancia["num_orders"]
    fila["num_lines"] = instancia["num_lines"]
    fila["toneladas_totales"] = toneladas_totales
    fila["media_toneladas"] = instancia["media_toneladas"]
    fila["desv_toneladas"] = instancia["desv_toneladas"]
    fila["min_load_open_line"] = instancia["min_load_open_line"]
    fila["max_load_open_line"] = instancia["max_load_open_line"]
    fila["penalizacion_linea_cerrada"] = instancia["penalizacion_linea_cerrada"]
    fila["prob_compatibilidad"] = instancia["prob_compatibilidad"]
    fila["prob_coste_cero"] = instancia["prob_coste_cero"]
    fila["prob_coste_inicio_cero"] = instancia["prob_coste_inicio_cero"]
    fila["use_seed"] = instancia["use_seed"]
    fila["seed_base"] = instancia["seed_base"]
    fila["seed_instancia"] = instancia["seed_instancia"]
    fila["match_instances_between_configs"] = instancia["match_instances_between_configs"]
    fila["estado_gurobi"] = estado_gurobi_excel(resultado)
    fila["obj_gurobi"] = resultado["obj_gurobi"]
    fila["gap_gurobi"] = resultado["mip_gap"]
    fila["tiempo_gurobi"] = resultado["runtime_gurobi"]
    fila["alns_ejecutado"] = resultado["alns_ejecutado"]
    fila["obj_alns"] = resultado["obj_alns"]
    fila["resultado_final"] = resultado_final_excel(resultado)
    fila["input_file"] = ruta_input(instancia["instancia_id"])
    fila["output_file"] = ruta_output(instancia["instancia_id"])

    if resultado["solucion_gurobi"] is not None:
        sol = resultado["solucion_gurobi"]
        toneladas = toneladas_asignadas(sol, weight)
        fila["lineas_abiertas_gurobi"] = numero_lineas_abiertas(sol)
        fila["lineas_cerradas_gurobi"] = lineas_cerradas(sol, lines)
        fila["toneladas_asignadas_gurobi"] = toneladas
        fila["toneladas_pendientes_gurobi"] = toneladas_totales - toneladas
        fila["porcentaje_toneladas_gurobi"] = 100 * toneladas / toneladas_totales if toneladas_totales > 0 else 0
        fila["orders_asignadas_gurobi"] = len(pedidos_asignados(sol))
        fila["orders_pendientes_gurobi"] = len(pedidos_pendientes(sol, orders))
        # Para Gurobi se usa como referencia el coste calculado directamente
        # desde los arcos x, porque no depende de la reconstrucción de la secuencia.
        fila["coste_setup_gurobi"] = resultado["coste_setup_gurobi_desde_x"]
        fila["coste_setup_gurobi_desde_x"] = resultado["coste_setup_gurobi_desde_x"]
        fila["coste_setup_gurobi_desde_secuencia"] = resultado["coste_setup_gurobi_desde_secuencia"]
        fila["diferencia_coste_setup_x_vs_secuencia"] = resultado["diferencia_coste_setup_x_vs_secuencia"]
        fila["secuencias_validas_gurobi"] = resultado["secuencias_validas_gurobi"]

    if resultado["solucion_alns"] is not None:
        sol = resultado["solucion_alns"]
        toneladas = toneladas_asignadas(sol, weight)
        fila["tiempo_alns"] = resultado["stats_alns"]["tiempo"]
        fila["iteraciones_alns"] = resultado["stats_alns"]["iteraciones"]
        fila["lineas_abiertas_alns"] = numero_lineas_abiertas(sol)
        fila["lineas_cerradas_alns"] = lineas_cerradas(sol, lines)
        fila["toneladas_asignadas_alns"] = toneladas
        fila["toneladas_pendientes_alns"] = toneladas_totales - toneladas
        fila["porcentaje_toneladas_alns"] = 100 * toneladas / toneladas_totales if toneladas_totales > 0 else 0
        fila["orders_asignadas_alns"] = len(pedidos_asignados(sol))
        fila["orders_pendientes_alns"] = len(pedidos_pendientes(sol, orders))
        fila["coste_setup_alns"] = coste_solucion(sol, coste, coste_inicio)

        obj_ini = resultado["stats_alns"]["objetivo_inicial"]
        obj_mejor = resultado["stats_alns"]["mejor_objetivo"]
        coste_ini = resultado["stats_alns"]["coste_setup_inicial"]
        coste_mejor = resultado["stats_alns"]["coste_setup_mejor"]

        fila["mejora_objetivo_alns"] = obj_ini - obj_mejor
        fila["mejora_porcentual_objetivo_alns"] = 100 * (obj_ini - obj_mejor) / obj_ini if abs(obj_ini) > 1e-9 else 0
        fila["mejora_coste_setup_alns"] = coste_ini - coste_mejor
        fila["mejora_porcentual_coste_setup_alns"] = 100 * (coste_ini - coste_mejor) / coste_ini if abs(coste_ini) > 1e-9 else 0

    return [fila[col] for col in COLUMNAS_EXCEL]


def colorear_resultado_final(ws, fila_excel):
    """Aplica color suave a la celda resultado_final."""
    col_idx = COLUMNAS_EXCEL.index("resultado_final") + 1
    celda = ws.cell(row=fila_excel, column=col_idx)
    color = COLORES_RESULTADO_FINAL.get(celda.value)

    if color is not None:
        celda.fill = PatternFill("solid", fgColor=color)

    celda.font = Font(bold=True)
    celda.alignment = Alignment(horizontal="center")


def actualizar_excel(instancia, resultado):
    """Añade una fila al Excel resumen y guarda el archivo tras cada instancia."""
    inicializar_excel()

    wb = load_workbook(EXCEL_RESUMEN)
    ws = wb["Resultados"]
    ws.append(construir_fila_excel(instancia, resultado))

    fila_excel = ws.max_row
    colorear_resultado_final(ws, fila_excel)

    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    wb.save(EXCEL_RESUMEN)



# ============================================================
# 10. RESUMEN Y GRÁFICOS DEL EXPERIMENTO 1
# ============================================================

def _valores_numericos(valores):
    """Filtra valores numéricos válidos para calcular medias y medianas."""
    filtrados = []
    for v in valores:
        if isinstance(v, (int, float)):
            filtrados.append(v)
    return filtrados


def _media(valores):
    valores = _valores_numericos(valores)
    if not valores:
        return None
    return sum(valores) / len(valores)


def _mediana(valores):
    valores = sorted(_valores_numericos(valores))
    n = len(valores)
    if n == 0:
        return None
    mitad = n // 2
    if n % 2 == 1:
        return valores[mitad]
    return (valores[mitad - 1] + valores[mitad]) / 2


def _porcentaje(parte, total):
    return 100 * parte / total if total else 0


def crear_resumen_y_graficos_excel():
    """
    Crea dos hojas adicionales al finalizar el experimento:
    - Resumen_por_tamano: métricas agregadas por número de orders.
    - Graficos: gráficos principales y box plots de tiempo y gap.

    La hoja Resultados no se modifica en estructura ni en columnas.
    """
    if not os.path.exists(EXCEL_RESUMEN):
        return

    wb = load_workbook(EXCEL_RESUMEN)
    if "Resultados" not in wb.sheetnames:
        wb.save(EXCEL_RESUMEN)
        return

    ws_res = wb["Resultados"]

    # Elimina hojas agregadas antiguas si se vuelve a ejecutar la función.
    for sheet_name in ["Resumen_por_tamano", "Graficos"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    headers = [cell.value for cell in ws_res[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}

    required = ["num_orders", "estado_gurobi", "gap_gurobi", "tiempo_gurobi"]
    for name in required:
        if name not in col:
            wb.save(EXCEL_RESUMEN)
            return

    datos = {}
    for row in range(2, ws_res.max_row + 1):
        n_orders = ws_res.cell(row=row, column=col["num_orders"]).value
        if n_orders is None:
            continue

        estado = ws_res.cell(row=row, column=col["estado_gurobi"]).value
        gap = ws_res.cell(row=row, column=col["gap_gurobi"]).value
        tiempo = ws_res.cell(row=row, column=col["tiempo_gurobi"]).value

        datos.setdefault(n_orders, {
            "estados": [],
            "gaps": [],
            "tiempos": [],
        })
        datos[n_orders]["estados"].append(estado)
        datos[n_orders]["gaps"].append(gap)
        datos[n_orders]["tiempos"].append(tiempo)

    ws_sum = wb.create_sheet("Resumen_por_tamano")
    resumen_headers = [
        "num_orders",
        "num_instancias",
        "num_optimas",
        "num_factibles_no_optimas",
        "num_sin_solucion_factible",
        "num_con_solucion_factible",
        "porcentaje_optimas",
        "porcentaje_factibles_no_optimas",
        "porcentaje_sin_solucion_factible",
        "porcentaje_con_solucion_factible",
        "gap_medio",
        "gap_mediana",
        "gap_minimo",
        "gap_maximo",
        "tiempo_medio",
        "tiempo_mediana",
        "tiempo_minimo",
        "tiempo_maximo",
    ]
    ws_sum.append(resumen_headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws_sum[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for n_orders in sorted(datos):
        estados = datos[n_orders]["estados"]
        gaps = _valores_numericos(datos[n_orders]["gaps"])
        tiempos = _valores_numericos(datos[n_orders]["tiempos"])

        total = len(estados)
        num_opt = sum(1 for e in estados if e == "Solución óptima")
        num_factible_no_opt = sum(1 for e in estados if e == "Solución factible")
        num_sin = sum(1 for e in estados if e == "Sin solución factible")
        num_factible = num_opt + num_factible_no_opt

        ws_sum.append([
            n_orders,
            total,
            num_opt,
            num_factible_no_opt,
            num_sin,
            num_factible,
            _porcentaje(num_opt, total),
            _porcentaje(num_factible_no_opt, total),
            _porcentaje(num_sin, total),
            _porcentaje(num_factible, total),
            _media(gaps),
            _mediana(gaps),
            min(gaps) if gaps else None,
            max(gaps) if gaps else None,
            _media(tiempos),
            _mediana(tiempos),
            min(tiempos) if tiempos else None,
            max(tiempos) if tiempos else None,
        ])

    ws_sum.freeze_panes = "A2"
    ws_sum.auto_filter.ref = ws_sum.dimensions

    for row in ws_sum.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00"

    for col_cells in ws_sum.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_sum.column_dimensions[col_letter].width = min(max_len + 2, 35)

    ws_graf = wb.create_sheet("Graficos")
    ws_graf["A1"] = "Gráficos del experimento 1: escalabilidad del modelo exacto"
    ws_graf["A1"].font = Font(bold=True, size=14)

    max_row = ws_sum.max_row

    # 1) Distribución de estados por tamaño.
    chart_estado = BarChart()
    chart_estado.type = "bar"
    chart_estado.style = 10
    chart_estado.title = "Distribución de estados por tamaño"
    chart_estado.y_axis.title = "% de instancias"
    chart_estado.x_axis.title = "Número de orders"
    chart_estado.grouping = "stacked"
    chart_estado.overlap = 100
    data = Reference(ws_sum, min_col=7, max_col=9, min_row=1, max_row=max_row)
    cats = Reference(ws_sum, min_col=1, min_row=2, max_row=max_row)
    chart_estado.add_data(data, titles_from_data=True)
    chart_estado.set_categories(cats)
    chart_estado.height = 9
    chart_estado.width = 18
    chart_estado.dataLabels = DataLabelList()
    chart_estado.dataLabels.showVal = False
    ws_graf.add_chart(chart_estado, "A3")

    # 2) Porcentaje de óptimas.
    chart_opt = LineChart()
    chart_opt.title = "% de instancias óptimas por tamaño"
    chart_opt.y_axis.title = "% óptimas"
    chart_opt.x_axis.title = "Número de orders"
    data = Reference(ws_sum, min_col=7, min_row=1, max_row=max_row)
    chart_opt.add_data(data, titles_from_data=True)
    chart_opt.set_categories(cats)
    chart_opt.height = 8
    chart_opt.width = 18
    ws_graf.add_chart(chart_opt, "J3")

    # 3) Porcentaje con solución factible.
    chart_fact = LineChart()
    chart_fact.title = "% de instancias con solución factible por tamaño"
    chart_fact.y_axis.title = "% con solución factible"
    chart_fact.x_axis.title = "Número de orders"
    data = Reference(ws_sum, min_col=10, min_row=1, max_row=max_row)
    chart_fact.add_data(data, titles_from_data=True)
    chart_fact.set_categories(cats)
    chart_fact.height = 8
    chart_fact.width = 18
    ws_graf.add_chart(chart_fact, "A22")

    # 4) Gap medio.
    chart_gap = LineChart()
    chart_gap.title = "Gap medio por tamaño"
    chart_gap.y_axis.title = "Gap medio"
    chart_gap.x_axis.title = "Número de orders"
    data = Reference(ws_sum, min_col=11, min_row=1, max_row=max_row)
    chart_gap.add_data(data, titles_from_data=True)
    chart_gap.set_categories(cats)
    chart_gap.height = 8
    chart_gap.width = 18
    ws_graf.add_chart(chart_gap, "J22")

    # 5) Tiempo medio de resolución.
    chart_tiempo = LineChart()
    chart_tiempo.title = "Tiempo medio de resolución por tamaño"
    chart_tiempo.y_axis.title = "Tiempo medio (s)"
    chart_tiempo.x_axis.title = "Número de orders"
    data = Reference(ws_sum, min_col=15, min_row=1, max_row=max_row)
    chart_tiempo.add_data(data, titles_from_data=True)
    chart_tiempo.set_categories(cats)
    chart_tiempo.height = 8
    chart_tiempo.width = 18
    ws_graf.add_chart(chart_tiempo, "A41")

    # Box plots con matplotlib, insertados como imágenes.
    if plt is not None and ExcelImage is not None and datos:
        try:
            labels = [str(n) for n in sorted(datos)]
            tiempo_series = [
                _valores_numericos(datos[n]["tiempos"])
                for n in sorted(datos)
            ]
            gap_series = [
                _valores_numericos(datos[n]["gaps"])
                for n in sorted(datos)
            ]

            # Box plot tiempo.
            plt.figure(figsize=(12, 5))
            plt.boxplot(tiempo_series, labels=labels, showfliers=True)
            plt.xlabel("Número de orders")
            plt.ylabel("Tiempo Gurobi (s)")
            plt.title("Distribución del tiempo de resolución por tamaño")
            plt.xticks(rotation=45)
            plt.tight_layout()
            ruta_box_tiempo = os.path.join(CARPETA_RESULTADOS, "boxplot_tiempo_gurobi.png")
            plt.savefig(ruta_box_tiempo, dpi=160)
            plt.close()

            img_tiempo = ExcelImage(ruta_box_tiempo)
            img_tiempo.width = 720
            img_tiempo.height = 300
            ws_graf.add_image(img_tiempo, "J41")

            # Box plot gap.
            plt.figure(figsize=(12, 5))
            plt.boxplot(gap_series, labels=labels, showfliers=True)
            plt.xlabel("Número de orders")
            plt.ylabel("Gap Gurobi")
            plt.title("Distribución del gap final por tamaño")
            plt.xticks(rotation=45)
            plt.tight_layout()
            ruta_box_gap = os.path.join(CARPETA_RESULTADOS, "boxplot_gap_gurobi.png")
            plt.savefig(ruta_box_gap, dpi=160)
            plt.close()

            img_gap = ExcelImage(ruta_box_gap)
            img_gap.width = 720
            img_gap.height = 300
            ws_graf.add_image(img_gap, "A60")
        except Exception as exc:
            ws_graf["A60"] = f"No se pudieron generar los box plots automáticamente: {exc}"
    else:
        ws_graf["A60"] = "Box plots no generados: matplotlib u openpyxl Image no disponibles."

    wb.save(EXCEL_RESUMEN)


# ============================================================
# 10. EJECUCIÓN PRINCIPAL
# ============================================================

def main():
    """Ejecuta el experimento completo instancia a instancia."""
    preparar_carpetas()
    inicializar_excel()

    total_instancias = total_instancias_experimento()
    instancia_id = 0

    for indice_config, config in enumerate(CONFIGURACIONES, start=1):
        for indice_local in range(1, config["num_instancias"] + 1):
            instancia_id += 1

            instancia = generar_instancia(instancia_id, indice_config, indice_local, config)
            guardar_input_generado(instancia)
            resultado = resolver_instancia(instancia, total_instancias)
            guardar_output_instancia(instancia, resultado)
            actualizar_excel(instancia, resultado)

            if PRINT_PROGRESS:
                print("\n--------------------------------------------------")
                print(f"Instancia {instancia_id}/{total_instancias} completada")
                print(f"Configuración: {instancia['nombre_configuracion']} | local {indice_local}/{config['num_instancias']}")
                print(f"Input: {ruta_input(instancia_id)}")
                print(f"Output: {ruta_output(instancia_id)}")
                print(f"Excel: {EXCEL_RESUMEN}")
                print("--------------------------------------------------\n")

    crear_resumen_y_graficos_excel()

    if PRINT_PROGRESS:
        print("\n==================================================")
        print("RESUMEN Y GRÁFICOS DEL EXPERIMENTO GENERADOS")
        print(f"Excel: {EXCEL_RESUMEN}")
        print("==================================================\n")


if __name__ == "__main__":
    main()
